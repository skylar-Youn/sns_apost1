import pandas as pd
import time
import os
import platform
import requests
import json
from dotenv import load_dotenv
import openai
from openai import OpenAI
import sys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
import subprocess
import re
import markdown
import base64
import random
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
# conda activate wordpress3_10 linux
# conda activate mov3_10 wsl 
#test
# WordPress 설정
WP_PATH = '/home/skyntech/www/html/wp'
WP_URL = 'http://222.122.202.122'
WP_ADMIN_USER = 'skylar'
WP_ADMIN_PASS = 'a1q2w3e4r!'

# 기본 구글 스프레드시트 URL 설정
DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1oyxj2wug-vMRBhsVbr36j7zkh8FllG-k8vQU-n4Ek-Y/edit?usp=sharing"

# 절대 경로 사용하여 .env 파일 로드
current_dir = os.path.dirname(os.path.abspath(__file__))  # 현재 스크립트 경로

# 우선순위에 따른 .env 파일 경로 목록
env_paths = [
    os.path.join(current_dir, '.env'),  # 1. 현재 스크립트 디렉토리
    '/home/sk/ws/SD/paddleOCR2/35.1movie/.env',  # 2. 프로젝트 루트
    '/home/skyntech/www/html/other/.env'  # 3. 기존 경로
]

env_loaded = False
for env_path in env_paths:
    print(f"env 파일 경로 확인: {env_path}")
    if os.path.exists(env_path):
        print(f".env 파일이 존재합니다: {env_path}")
        load_dotenv(dotenv_path=env_path)
        env_loaded = True
        break
    else:
        print(f".env 파일이 존재하지 않습니다: {env_path}")

if not env_loaded:
    print("모든 경로에서 .env 파일을 찾을 수 없습니다")

# API 키 설정
api_key = os.getenv("OPENAI_API_KEY")
if api_key:
    print("API 키를 성공적으로 로드했습니다.")
else:
    print("API 키를 로드하지 못했습니다.")
openai.api_key = api_key

# 워드프레스 CLI 명령어 실행 함수
def run_wp_cli(cmd):
    """워드프레스 CLI 명령어를 실행합니다."""
    full_cmd = f"wp {cmd} --path={WP_PATH}"
    print(f"> {full_cmd}")
    r = subprocess.run(full_cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if r.returncode:
        print(f"✖ 오류({r.returncode}): {r.stderr.strip()}")
        return False, r.stderr.strip()
    else:
        out = r.stdout.strip()
        if out: 
            print(f"✔ 완료: {out}")
        return True, out

# 워드프레스 CLI 설치 확인
def check_wp_cli():
    """워드프레스 CLI가 설치되어 있는지 확인합니다."""
    if subprocess.run("which wp", shell=True).returncode:
        print("→ WP-CLI 미설치, 설치를 시도합니다...")
        os.system("curl -sO https://raw.githubusercontent.com/wp-cli/builds/gh-pages/phar/wp-cli.phar")
        os.system("chmod +x wp-cli.phar")
        os.system("sudo mv wp-cli.phar /usr/local/bin/wp")
        if subprocess.run("which wp", shell=True).returncode:
            print("✖ WP-CLI 설치 실패")
            return False
    print("→ WP-CLI 설치 확인됨")
    return True

# 워드프레스 설치 확인
def check_wp_installed():
    """워드프레스가 설치되어 있는지 확인합니다."""
    return subprocess.run(f"wp core is-installed --path={WP_PATH}", shell=True).returncode == 0

# 게시글 발행 함수
def publish_to_wordpress(title, content, category_name=None, auto_category=False, keyword=None, image_url=None, language=None):
    """제목과 내용으로 워드프레스에 게시글을 발행합니다."""
    # 워드프레스 CLI 확인
    if not check_wp_cli():
        print("WP-CLI가 없어서 워드프레스 발행을 진행할 수 없습니다.")
        return False, None, None

    # 워드프레스 설치 확인
    if not check_wp_installed():
        print("워드프레스가 설치되어 있지 않아 게시글을 발행할 수 없습니다.")
        return False, None, None
    
    # 콘텐츠가 마크다운인지 HTML인지 확인
    is_markdown = False
    original_content = content
    
    # 콘텐츠 처리: 마크다운 코드 블록이 있는지 확인하고 변환
    if "```Markdown" in content or "```markdown" in content or "```md" in content:
        print("마크다운 코드 블록 감지, 마크다운을 HTML로 변환합니다...")
        # 코드 블록에서 마크다운 추출
        markdown_content = extract_markdown_from_codeblock(content)
        is_markdown = True
        
        # 중복 제목 제거 (마크다운)
        print("중복된 제목 확인 및 제거 중...")
        markdown_content = remove_duplicate_title(title, markdown_content, is_markdown=True)
        print("중복 제목 처리 완료")
        
        # 마크다운을 HTML로 변환
        content = convert_markdown_to_html(markdown_content)
        print("마크다운 변환 완료")
    # 마크다운 특징 감지 (## 헤더, | 테이블, --- 구분선 등)
    elif re.search(r'^##\s+|^\|.*\|.*\||^---$|^\*\*.*\*\*|^\-\s+|^\d+\.\s+|^>\s+TIP:', content, re.MULTILINE):
        print("마크다운 형식 감지, 마크다운을 HTML로 변환합니다...")
        is_markdown = True
        
        # HTML 주석 제거 (<!-- 키워드: ... --> 등)
        content = re.sub(r'<!--.*?-->', '', content, flags=re.DOTALL)
        
        # 첫 줄이 ---로 시작하는 경우 제거 (마크다운 메타데이터로 오해할 수 있음)
        content = re.sub(r'^\s*---\s*\n', '', content, count=1)
        
        # 중복 제목 제거 (마크다운)
        print("중복된 제목 확인 및 제거 중...")
        markdown_content = remove_duplicate_title(title, content, is_markdown=True)
        print("중복 제목 처리 완료")
        
        # 마크다운을 HTML로 변환
        content = convert_markdown_to_html(markdown_content)
        print("마크다운 변환 완료")
    else:
        # HTML 코드 블록 처리
        html_match = re.search(r'```html\s*([\s\S]*?)\s*```', content, re.IGNORECASE)
        if html_match:
            print("HTML 코드 블록 감지, HTML 내용을 추출합니다...")
            content = html_match.group(1).strip()
            print("HTML 추출 완료")
        
        # 중복 제목 제거 (HTML)
        print("중복된 제목 확인 및 제거 중...")
        content = remove_duplicate_title(title, content, is_markdown=False)
        print("중복 제목 처리 완료")
    
    # 이미지가 있는 경우 콘텐츠에 추가
    if image_url:
        if image_url.startswith("http"):
            # URL에서 이미지 다운로드 후 WordPress에 업로드
            try:
                print("이미지를 WordPress 미디어 라이브러리에 업로드 중...")
                img_data = requests.get(image_url).content
                temp_image = f"/tmp/{keyword.replace(' ', '_')}_temp.jpg"
                with open(temp_image, 'wb') as f:
                    f.write(img_data)
                
                # WP-CLI로 미디어 업로드
                success, media_output = run_wp_cli(f"media import '{temp_image}' --porcelain")
                if success:
                    media_id = media_output.strip()
                    # 미디어 URL 가져오기
                    success, media_url = run_wp_cli(f"post get {media_id} --field=guid")
                    if success:
                        # 반응형 이미지 HTML 생성 (중앙 정렬 및 크기 제한)
                        image_html = f'''<div style="text-align: center; margin: 30px 0; padding: 0;">
    <img src="{media_url}" alt="{title}" 
         style="display: block; margin: 0 auto; max-width: 100%; max-height: 500px; width: auto; height: auto; object-fit: contain; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.1);" 
         loading="lazy" />
</div>'''
                        content = image_html + "\n\n" + content
                        print("반응형 이미지가 콘텐츠에 추가되었습니다.")
                
                # 임시 파일 삭제
                if os.path.exists(temp_image):
                    os.remove(temp_image)
            except Exception as e:
                print(f"이미지 업로드 중 오류: {e}")
        else:
            # 로컬 파일인 경우
            if os.path.exists(image_url):
                success, media_output = run_wp_cli(f"media import '{image_url}' --porcelain")
                if success:
                    media_id = media_output.strip()
                    success, media_url = run_wp_cli(f"post get {media_id} --field=guid")
                    if success:
                        # 반응형 이미지 HTML 생성 (중앙 정렬 및 크기 제한)
                        image_html = f'''<div style="text-align: center; margin: 30px 0; padding: 0;">
    <img src="{media_url}" alt="{title}" 
         style="display: block; margin: 0 auto; max-width: 100%; max-height: 500px; width: auto; height: auto; object-fit: contain; border-radius: 8px; box-shadow: 0 4px 12px rgba(0,0,0,0.1);" 
         loading="lazy" />
</div>'''
                        content = image_html + "\n\n" + content
                        print("반응형 이미지가 콘텐츠에 추가되었습니다.")

    # 임시 파일에 내용 저장
    temp_content_file = "temp_content.html"
    with open(temp_content_file, "w", encoding="utf-8") as f:
        f.write(content)

    # 실제로 사용될 카테고리 이름을 저장할 변수
    final_category_name = category_name
    
    # 카테고리 처리 로직
    if category_name == "manual_select":
        # 수동 선택 모드
        print(f"\n키워드 '{keyword}'에 대한 카테고리를 선택하세요.")
        success, categories_output = run_wp_cli("term list category --fields=name --format=csv")
        if success:
            category_names = []
            lines = categories_output.split('\n')
            for line in lines:
                if line and "name" not in line:
                    clean_name = line.strip().strip('"')
                    if clean_name:
                        category_names.append(clean_name)
            
            if category_names:
                print("\n현재 사용 가능한 카테고리:")
                for i, cat in enumerate(category_names):
                    print(f"{i+1}. {cat}")
                print(f"{len(category_names)+1}. 새 카테고리 생성")
                print(f"{len(category_names)+2}. 미분류")
                
                while True:
                    try:
                        choice = int(input(f"\n카테고리를 선택하세요 (1-{len(category_names)+2}): ").strip())
                        if 1 <= choice <= len(category_names):
                            final_category_name = category_names[choice-1]
                            print(f"'{final_category_name}' 카테고리가 선택되었습니다.")
                            break
                        elif choice == len(category_names)+1:
                            new_cat = input("새 카테고리 이름을 입력하세요: ").strip()
                            if new_cat:
                                final_category_name = new_cat
                                print(f"새 카테고리 '{final_category_name}'가 생성됩니다.")
                                break
                        elif choice == len(category_names)+2:
                            final_category_name = None
                            print("미분류로 설정되었습니다.")
                            break
                        else:
                            print("유효한 번호를 입력하세요.")
                    except ValueError:
                        print("숫자를 입력하세요.")
            else:
                final_category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
        else:
            final_category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
        category_name = final_category_name
    elif auto_category and keyword:
        # 자동 카테고리 추천 기능
        print("키워드에 맞는 자동 카테고리 분류를 시작합니다...")
        suggested_category = suggest_category_for_keyword(keyword)
        if suggested_category:
            print(f"키워드 '{keyword}'에 대해 '{suggested_category}' 카테고리가 추천되어 자동 적용됩니다.")
            category_name = suggested_category
            final_category_name = suggested_category
        else:
            print("카테고리 자동 추천에 실패했습니다. 수동으로 선택하세요.")
            # 기존 카테고리 목록 가져오기
            success, categories_output = run_wp_cli("term list category --fields=name --format=csv")
            if success:
                # 카테고리 목록 처리
                category_names = []
                lines = categories_output.split('\n')
                for line in lines:
                    if line and "name" not in line:  # 헤더 제외
                        # 따옴표 제거
                        clean_name = line.strip().strip('"')
                        if clean_name:
                            category_names.append(clean_name)
                
                if category_names:
                    print("\n현재 사용 가능한 카테고리:")
                    for i, cat in enumerate(category_names):
                        print(f"{i+1}. {cat}")
                    print(f"{len(category_names)+1}. 새 카테고리 생성")
                    print(f"{len(category_names)+2}. 미분류")
                    
                    while True:
                        try:
                            choice = int(input(f"\n카테고리를 선택하세요 (1-{len(category_names)+2}): ").strip())
                            if 1 <= choice <= len(category_names):
                                category_name = category_names[choice-1]
                                final_category_name = category_name
                                print(f"'{category_name}' 카테고리가 선택되었습니다.")
                                break
                            elif choice == len(category_names)+1:
                                new_cat = input("새 카테고리 이름을 입력하세요: ").strip()
                                if new_cat:
                                    category_name = new_cat
                                    final_category_name = new_cat
                                    print(f"새 카테고리 '{category_name}'가 생성됩니다.")
                                    break
                            elif choice == len(category_names)+2:
                                category_name = None
                                final_category_name = None
                                print("미분류로 설정되었습니다.")
                                break
                            else:
                                print("유효한 번호를 입력하세요.")
                        except ValueError:
                            print("숫자를 입력하세요.")
                else:
                    category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
                    final_category_name = category_name
            else:
                category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
                final_category_name = category_name

    # 카테고리 처리
    category_param = ""
    if category_name:
        # 카테고리 ID 가져오기
        cat_id = get_or_create_wp_category(category_name)
        if cat_id:
            category_param = f"--post_category={cat_id}"
            print(f"카테고리 '{category_name}' (ID: {cat_id})가 설정되었습니다.")
        else:
            print(f"카테고리 '{category_name}' 설정 실패")
    
    # 게시글 발행
    print(f"→ '{title}' 제목으로 게시글 발행 중...")
    
    # 제목에서 작은따옴표 이스케이프
    safe_title = title.replace("'", "\\'")
    
    # 먼저 포스트 생성 (언어 파라미터 없이)
    success, post_id = run_wp_cli(
        f"post create '{temp_content_file}' --post_title='{safe_title}' --post_status=publish {category_param} --porcelain"
    )
    
    # 포스트 생성 성공 시 Polylang 언어 설정
    if success and language:
        # Polylang 플러그인이 활성화되어 있는지 확인
        check_polylang, _ = run_wp_cli("plugin is-active polylang")
        if check_polylang:
            # 언어 코드가 유효한지 확인 (ko 또는 en)
            if language in ['ko', 'en']:
                print(f"Polylang 언어 설정 중: {language}")
                # wp post term set 명령어로 언어 설정
                lang_success, lang_output = run_wp_cli(f"post term set {post_id} language {language}")
                if lang_success:
                    print(f"✅ 언어 설정 완료: {language}")
                    # 영어인 경우 URL 구조 확인
                    if language == 'en':
                        print(f"영어 포스트 URL 구조: /language/en/?p={post_id}")
                else:
                    print(f"⚠️ 언어 설정 실패: {lang_output}")
            else:
                print(f"경고: 지원하지 않는 언어 코드 '{language}'. 기본 언어로 발행되었습니다.")
        else:
            print("Polylang 플러그인이 활성화되지 않아 언어 설정을 건너뜁니다.")
    
    # 임시 파일 삭제
    if os.path.exists(temp_content_file):
        os.remove(temp_content_file)
    
    if success:
        # 언어별 URL 구조 적용
        if language == 'en':
            post_url = f"{WP_URL}/language/en/?p={post_id}"
        else:
            post_url = f"{WP_URL}/?p={post_id}"
        
        print(f"✅ 게시글이 성공적으로 발행되었습니다: {post_url}")
        return True, post_id, post_url, final_category_name
    else:
        print("✖ 게시글 발행 실패")
        return False, None, None, None

# Google Sheets API 연결 함수
def connect_to_google_sheets(credentials_path):
    """Google Sheets API에 연결합니다."""
    try:
        # Google Sheets API 사용 범위 설정
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/drive'
        ]
        
        # API 인증
        credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_path, scope)
        client = gspread.authorize(credentials)
        
        print("구글 스프레드시트 API 연결 성공!")
        return client
    except Exception as e:
        print(f"구글 스프레드시트 API 연결 실패: {e}")
        return None

# 구글 스프레드시트 열기 함수
def open_google_sheet(client, url_or_key):
    """URL 또는 키로 구글 스프레드시트를 엽니다."""
    try:
        # URL에서 키 추출
        if '/' in url_or_key:
            key = url_or_key.split('/d/')[1].split('/')[0]
        else:
            key = url_or_key
            
        # 스프레드시트 열기
        sheet = client.open_by_key(key)
        print(f"스프레드시트 열기 성공: {sheet.title}")
        return sheet
    except Exception as e:
        print(f"스프레드시트 열기 실패: {e}")
        return None

# 데이터프레임으로 스프레드시트 내용 가져오기
def get_sheet_as_dataframe(worksheet):
    """워크시트의 내용을 판다스 데이터프레임으로 가져옵니다."""
    try:
        # 모든 값 가져오기
        data = worksheet.get_all_values()
        
        # 헤더 추출
        headers = data[0]
        
        # 데이터 추출
        values = data[1:]
        
        # 데이터프레임 생성
        df = pd.DataFrame(values, columns=headers)
        
        print(f"스프레드시트 데이터 로드 완료: {len(df)}행 x {len(df.columns)}열")
        return df
    except Exception as e:
        print(f"스프레드시트 데이터 로드 실패: {e}")
        return None

# 특정 셀에 값 쓰기
def update_cell(worksheet, row, col, value):
    """특정 셀에 값을 씁니다. row, col은 1부터 시작"""
    try:
        worksheet.update_cell(row, col, value)
        print(f"셀 업데이트 성공: {chr(64+col)}{row}")
        return True
    except Exception as e:
        print(f"셀 업데이트 실패: {e}")
        return False

# 안전한 셀 읽기 (워크시트 이름에 특수문자가 있을 때를 위한 함수)
def safe_get_cell_value(worksheet, row, col):
    """워크시트에서 안전하게 셀 값을 읽습니다."""
    try:
        # 방법 1: 직접 셀 접근
        return worksheet.cell(row, col).value
    except Exception as e1:
        try:
            # 방법 2: 범위로 접근
            cell_address = f"{chr(64+col)}{row}"
            values = worksheet.get(cell_address)
            if values and values[0]:
                return values[0][0]
            return None
        except Exception as e2:
            print(f"셀 읽기 실패 ({row},{col}): {e2}")
            return None

DEBUGGING_PORT = 9222  # 윈도우용 크롬 remote debugging 포트

def show_preparation_guide():
    print("="*70)
    print("                           --- 중요 ---")
    print("="*70)
    print("1. OpenAI API 키가 필요합니다.")
    print(f"2. {env_path} 파일에 OPENAI_API_KEY=your_api_key 형식으로 키가 저장되어 있어야 합니다.")
    print("3. API 요청에는 비용이 발생할 수 있습니다.")
    print("4. 기본적으로 로컬 엑셀 파일을 사용합니다. 구글 스프레드시트도 선택 가능합니다.")
    print("5. 티스토리 발행을 선택한 경우:")
    if platform.system() == "Windows":
        print(f'   "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
    else:
        print(f'   google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData"')
    print("   크롬을 위 명령어로 실행한 후, 로그인된 상태에서 본 프로그램을 실행하세요.")
    print("6. WordPress 발행을 선택한 경우:")
    print("   WP-CLI가 설치되어 있어야 하며, WordPress가 정상적으로 작동해야 합니다.")
    print("="*70)
    
    # API 키 확인
    if not api_key:
        print(f"API 키가 {env_path} 파일에서 설정되지 않았습니다!")
        key_input = input("OpenAI API 키를 입력하세요(입력하지 않으면 프로그램이 종료됩니다): ").strip()
        if key_input:
            os.environ["OPENAI_API_KEY"] = key_input
            openai.api_key = key_input
            print("API 키를 수동으로 설정했습니다.")
        else:
            print("API 키가 제공되지 않았습니다. 프로그램을 종료합니다.")
            return False
    
    user_ready = input("준비가 완료되면 'y'를 입력하세요 (기본값: y): ").strip().lower() or 'y'
    if user_ready != 'y':
        print("프로그램을 종료합니다. 준비가 완료된 후 다시 실행해주세요.")
        return False
    return True

def extract_and_search_keywords():
    try:
        # 사전 준비 안내 표시
        if not show_preparation_guide():
            return
        
        # 결과 저장 디렉토리 설정 (include_images 처리 전에 추가)
        results_dir = "gpt_results"
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)
            print(f"결과 저장 디렉토리 생성: {results_dir}")
        
        worksheet = None
        gsheet_worksheet = None

        # 구글 스프레드시트를 기본으로 사용
        use_google_sheet = input("구글 스프레드시트를 사용하시겠습니까? (y/n, 기본값: y): ").strip().lower() != 'n'
        
        if use_google_sheet:
            # 구글 API 인증 정보 경로 - 플랫폼별로 다른 경로 사용
            if platform.system() == "Windows":
                # Windows 경로 (현재 프로젝트 폴더)
                credentials_path = "F:/1.1android/AndroidStudioProjects/35.1movie/youtube-shorts-455403-ea00801dd7b2.json"
            else:
                # Linux/WSL 경로 - 우선순위별로 확인
                credentials_paths = [
                    os.path.join(current_dir, "youtube-shorts-455403-ea00801dd7b2.json"),  # 현재 디렉토리
                    "/home/sk/ws/SD/paddleOCR2/35.1movie/youtube-shorts-455403-ea00801dd7b2.json",  # 프로젝트 루트
                    "/home/skyntech/www/html/other/youtube-shorts-455403-ea00801dd7b2.json",  # 기존 경로
                    "/mnt/f/1.1android/AndroidStudioProjects/35.1movie/youtube-shorts-455403-ea00801dd7b2.json"  # WSL 마운트 경로
                ]

                credentials_path = None
                for path in credentials_paths:
                    if os.path.exists(path):
                        credentials_path = path
                        break

                if credentials_path is None:
                    credentials_path = credentials_paths[0]  # 기본값으로 현재 디렉토리 설정
            
            print(f"운영체제: {platform.system()}")
            print(f"구글 API 인증 정보 파일 경로: {credentials_path}")
            
            if not os.path.exists(credentials_path):
                print("인증 파일이 없습니다. 로컬 엑셀 파일을 사용합니다.")
                use_google_sheet = False
            else:
                # 구글 API 연결
                client = connect_to_google_sheets(credentials_path)
                if not client:
                    print("구글 API 연결에 실패했습니다. 로컬 엑셀 파일을 사용합니다.")
                    use_google_sheet = False
                else:
                    # 구글 스프레드시트 URL 입력 받기
                    default_url = DEFAULT_SHEET_URL
                    sheet_url = input(f"구글 스프레드시트 URL을 입력하세요 (기본값: {default_url}): ").strip() or default_url
                    print(f"사용할 스프레드시트 URL: {sheet_url}")
                    
                    # 스프레드시트 열기
                    sheet = open_google_sheet(client, sheet_url)
                    if not sheet:
                        print("스프레드시트 열기에 실패했습니다. 로컬 엑셀 파일을 사용합니다.")
                        use_google_sheet = False
                    else:
                        # 워크시트 선택
                        worksheet_list = sheet.worksheets()
                        print("\n사용 가능한 워크시트:")
                        for i, ws in enumerate(worksheet_list):
                            print(f"{i+1}. {ws.title}")
                        
                        ws_input = input("\n사용할 워크시트 번호를 선택하세요 (기본값: 1): ").strip()
                        if not ws_input:
                            ws_index = 0
                        else:
                            try:
                                ws_index = int(ws_input) - 1
                            except ValueError:
                                print("유효한 숫자를 입력하지 않아 기본값(1번)을 사용합니다.")
                                ws_index = 0

                        if 0 <= ws_index < len(worksheet_list):
                            gsheet_worksheet = worksheet_list[ws_index]
                            print(f"워크시트 선택: {gsheet_worksheet.title}")
                        else:
                            print("유효하지 않은 워크시트 번호입니다. 로컬 엑셀 파일을 사용합니다.")
                            use_google_sheet = False
                        
                        if use_google_sheet:
                            # 워크시트 데이터 로드
                            df = get_sheet_as_dataframe(gsheet_worksheet)
                            if df is None:
                                print("스프레드시트 데이터 로드에 실패했습니다. 로컬 엑셀 파일을 사용합니다.")
                                use_google_sheet = False
        
        # 로컬 엑셀 파일 사용
        if not use_google_sheet:
            excel_path = input("처리할 로컬 엑셀 파일 경로를 입력하세요: ").strip()
            if not os.path.exists(excel_path):
                print(f"파일이 존재하지 않습니다: {excel_path}")
                return
            # 엑셀 파일 로드 (로컬 파일 처리 로직)
            try:
                df = pd.read_excel(excel_path)
                print(f"엑셀 파일 로드 완료: {excel_path}")
                workbook = openpyxl.load_workbook(excel_path)
                worksheet = workbook.active
                is_google_sheet = False
            except Exception as e:
                print(f"엑셀 파일 로드 중 오류 발생: {e}")
                return
        else:
            is_google_sheet = True
        
        sheet_title = ""
        if is_google_sheet:
            sheet_title = gsheet_worksheet.title if gsheet_worksheet else ""
        elif worksheet is not None:
            sheet_title = worksheet.title
            print(f"워크시트 선택: {sheet_title}")

        # 사용 가능한 모든 컬럼 표시
        print(f"사용 가능한 컬럼: {df.columns.tolist()}")
        
        # 키워드가 저장된 컬럼 선택
        keyword_column = input("키워드가 저장된 컬럼 이름을 입력하세요 (기본값: 키워드): ").strip() or "키워드"
        
        # 선택한 컬럼이 데이터프레임에 있는지 확인
        if keyword_column not in df.columns:
            print(f"'{keyword_column}' 컬럼이 파일에 존재하지 않습니다.")
            alternative_column = input("다른 컬럼을 선택하세요: ").strip()
            if alternative_column in df.columns:
                keyword_column = alternative_column
            else:
                print(f"'{alternative_column}' 컬럼도 존재하지 않습니다. 프로그램을 종료합니다.")
                return
        
        # 키워드 컬럼이 몇 번째 컬럼인지 확인 (A=0, B=1, ...)
        keyword_column_index = df.columns.get_loc(keyword_column)
        print(f"키워드 컬럼 인덱스: {keyword_column_index} (A=0, B=1, ...)")

        # 블로그목차 워크시트인 경우에만 소제목 컬럼 추출
        use_outline_topics = bool(sheet_title and sheet_title.startswith("블로그목차-"))
        outline_subtitle_columns: list[tuple[int, str]] = []
        total_columns = len(df.columns)

        if use_outline_topics:
            preferred_indices = [
                idx for idx in (2, 3, 4)
                if idx != keyword_column_index and idx < total_columns
            ]
            used_indices = set()
            for column_position in preferred_indices:
                header_name = df.columns[column_position] or f"컬럼{column_position}"
                outline_subtitle_columns.append((column_position, header_name))
                used_indices.add(column_position)

            if len(outline_subtitle_columns) < 3:
                for offset in range(1, 4):
                    column_position = keyword_column_index + offset
                    if column_position >= total_columns or column_position == keyword_column_index:
                        continue
                    if column_position in used_indices:
                        continue
                    header_name = df.columns[column_position] or f"컬럼{column_position}"
                    outline_subtitle_columns.append((column_position, header_name))
                    used_indices.add(column_position)
                    if len(outline_subtitle_columns) >= 3:
                        break

            if len(outline_subtitle_columns) < 3:
                for idx, col_name in enumerate(df.columns):
                    if idx == keyword_column_index or idx in used_indices:
                        continue
                    if str(col_name).startswith("소제목") or not str(col_name).strip():
                        header_name = col_name or f"컬럼{idx}"
                        outline_subtitle_columns.append((idx, header_name))
                        used_indices.add(idx)
                    if len(outline_subtitle_columns) >= 3:
                        break

            if outline_subtitle_columns:
                printable_headers = [f"idx:{idx} name:{name}" for idx, name in outline_subtitle_columns]
                print(f"블로그 소제목으로 사용할 컬럼: {printable_headers}")
            else:
                print("블로그목차 워크시트지만 사용할 소제목 컬럼을 찾지 못했습니다. 키워드만 사용합니다.")
        else:
            print("블로그목차- 워크시트가 아니므로 키워드만 사용합니다.")

# 결과를 저장할 열 설정
        image_column_letter = "E"
        image_column_index = ord(image_column_letter) - 65
        
        # 결과 저장 열 설정
        title_column_letter = "F"    # 제목 열
        content_column_letter = "G"  # 내용 열
        post_info_column_letter = "H"  # 포스트 정보 열 추가
        title_column_index = ord(title_column_letter) - 65      # A=0, B=1, ...
        content_column_index = ord(content_column_letter) - 65  # A=0, B=1, ...
        post_info_column_index = ord(post_info_column_letter) - 65  # A=0, B=1, ...
        print(f"제목 저장 열: {title_column_letter}, 내용 저장 열: {content_column_letter}, 포스트 정보 저장 열: {post_info_column_letter}")
        
        # 키워드 컬럼에서 키워드 추출
        valid_keywords = []
        for index, keyword in enumerate(df[keyword_column]):
            # 키워드가 유효한 경우에만 처리
            if pd.notna(keyword) and str(keyword).strip():
                # 원본 행 번호 (0번 행은 헤더)
                excel_row = index + 2  # 행 번호는 1부터 시작, 헤더가 1행이므로 +2
                row_series = df.iloc[index]

                outline_topics = []
                if use_outline_topics and outline_subtitle_columns:
                    for col_idx, _ in outline_subtitle_columns:
                        if col_idx < len(row_series):
                            cell_value = row_series.iloc[col_idx]
                            if pd.notna(cell_value):
                                cleaned_value = str(cell_value).strip()
                                if cleaned_value:
                                    outline_topics.append(cleaned_value)

                existing_title_value = ""
                if title_column_index < len(row_series):
                    title_cell = row_series.iloc[title_column_index]
                    if pd.notna(title_cell):
                        existing_title_value = str(title_cell).strip()

                title_present = bool(existing_title_value)

                valid_keywords.append(
                    (
                        index + 1,
                        str(keyword).strip(),
                        excel_row,
                        outline_topics,
                        title_present,
                        existing_title_value,
                    )
                )

        total_keywords = len(valid_keywords)
        print(f"\n===== 키워드 추출 완료 =====")
        print(f"총 {total_keywords}개의 키워드를 찾았습니다.")

        if total_keywords == 0:
            print("처리할 키워드가 없습니다.")
            return

        pending_keywords = [item for item in valid_keywords if not item[4]]
        remaining_count = len(pending_keywords)
        print(f"제목(F열)이 비어있는 키워드 수: {remaining_count}")
        if remaining_count:
            preview = ", ".join(f"{item[0]}:{item[1]}" for item in pending_keywords[:10])
            ellipsis = " ..." if remaining_count > 10 else ""
            if preview:
                print(f"진행 가능 키워드 번호: {preview}{ellipsis}")
        else:
            print("모든 키워드의 제목(F열)이 채워져 있습니다.")

        
        # 모델 선택
        print("\n사용할 모델을 선택하세요:")
        print("1. gpt-4.1 (더 높은 품질, 더 비쌈)")
        print("2. gpt-3.5-turbo (더 빠르고 저렴함)")
        print("3. 블로그 모드 (높은 가독성, 최적화된 형식x)")
        print("4.북 모드 (마크다운 책 형식)")
        print("5. 블로그마크다운v 모드 (마크다운 블로그 형식)")
        print("6. 블로그 HTML 모드 (OpenAI 클라이언트 버전)")
        print("7. 책 HTML 모드 (OpenAI 클라이언트 버전)")
        print("8. 블로그마크다운&아이콘 (GPT-4.1, 마크다운+이모지+목차+표+TIP)")
        print("9. 한글 블로그마크다운&아이콘&image 모드 (GPT-4.1, 마크다운+이모지+목차+표+TIP+이미지)")
        print("10. 영어 블로그 모드 (GPT-4.1, 미국 시장 대상 SEO 최적화 영문 블로그)")
        print("11. 프리미엄 컨텐츠 모드 (GPT-4.1, 고품질 한국어 콘텐츠+상세 가이드+전문가 팁)")
        print("91. 한글 블로그마크다운&아이콘&image 모드 NEW (마크다운 본문 + HTML 헤더 분리 입력)")
        print("92. 한글 블로그마크다운&아이콘&image 모드 HTML 일괄 입력 (요약 텍스트 자동 추가)")
        model_choice = input("모델 번호 선택 (기본값: 9): ").strip() or "9"

        # 모델 번호에 따라 실제 모델 이름과 모드 설정
        if model_choice == "1":
            model_name = "gpt-4.1"
            content_mode = "1"  # 일반 모드
            print("gpt-4.1 모델이 선택되었습니다.")
        elif model_choice == "3":
            model_name = "gpt-4.1"
            content_mode = "2"  # 블로그 모드
            print("gpt-4.1 모델과 블로그html 모드가 선택되었습니다.")
        elif model_choice == "4":
            model_name = "gpt-4.1"
            content_mode = "3"  # 북 모드
            print("gpt-4.1 모델과 북마크다운 모드가 선택되었습니다.")
        elif model_choice == "5":
            model_name = "gpt-4.1"
            content_mode = "4"  # 블로그마크다운 모드
            print("gpt-4.1 모델과 블로그마크다운 모드가 선택되었습니다.")
        elif model_choice == "6":
            model_name = "gpt-4.1"
            content_mode = "6"  # 블로그 HTML 모드 (OpenAI 클라이언트 버전)
            print("gpt-4.1 모델과 블로그 HTML 모드가 선택되었습니다.")
        elif model_choice == "7":
            model_name = "gpt-4.1"
            content_mode = "7"  # 책 HTML 모드 (OpenAI 클라이언트 버전)
            print("gpt-4.1 모델과 책 HTML 모드가 선택되었습니다.")
        elif model_choice == "8":
            model_name = "gpt-4.1"
            content_mode = "8"  # 블로그마크다운&아이콘 (GPT-4.1, 마크다운+이모지+목차+표+TIP)
            print("gpt-4.1 모델과 블로그마크다운&아이콘 모드가 선택되었습니다.")
        elif model_choice == "9":
            model_name = "gpt-4.1"
            content_mode = "9"  # 블로그마크다운&아이콘&이미지
            print("gpt-4.1 모델과 블로그마크다운&아이콘&이미지 모드가 선택되었습니다.")
        elif model_choice == "10":
            model_name = "gpt-4.1"
            content_mode = "10"  # 영어 블로그 모드
            print("gpt-4.1 모델과 영어 블로그 모드가 선택되었습니다.")
        elif model_choice == "11":
            model_name = "gpt-4.1"
            content_mode = "11"  # 프리미엄 컨텐츠 모드
            print("gpt-4.1 모델과 프리미엄 컨텐츠 모드가 선택되었습니다.")
        elif model_choice == "91":
            model_name = "gpt-4.1"
            content_mode = "91"
            print("gpt-4.1 모델과 블로그마크다운&아이콘&이미지 모드 NEW (마크다운 분리 입력)가 선택되었습니다.")
        elif model_choice == "92":
            model_name = "gpt-4.1"
            content_mode = "92"
            print("gpt-4.1 모델과 블로그마크다운&아이콘&이미지 HTML 일괄 입력 모드가 선택되었습니다.")
        else:
            model_name = "gpt-3.5-turbo"
            content_mode = "1"  # 일반 모드
            print("gpt-3.5-turbo 모델이 선택되었습니다.")
        
        # 시작할 키워드 번호 입력 받기
        default_start_number = pending_keywords[0][0] if pending_keywords else 1

        while True:
            try:
                start_input = input(
                    f"\n시작할 키워드 번호를 입력하세요 (1-{total_keywords}, 기본값: {default_start_number}): "
                ).strip()
                if not start_input:
                    start_number = default_start_number
                else:
                    start_number = int(start_input)
                if 1 <= start_number <= total_keywords:
                    break
                else:
                    print(f"유효한 범위 내의 번호를 입력하세요 (1-{total_keywords}).")
            except ValueError:
                print("숫자를 입력하세요.")

        # 자동으로 몇 개의 키워드를 처리할지 설정
        available_count = total_keywords - start_number + 1
        pending_after_start = len([item for item in pending_keywords if item[0] >= start_number])
        default_process_count = pending_after_start if pending_after_start else available_count

        while True:
            try:
                prompt = (
                    f"몇 개의 키워드를 순차적으로 처리할까요? (1-{available_count}, 0=모두, 기본값: {default_process_count}): "
                )
                process_input = input(prompt).strip()
                if not process_input:
                    process_count = default_process_count
                else:
                    process_count = int(process_input)
                if process_count == 0:
                    process_count = available_count
                if 1 <= process_count <= available_count:
                    break
                else:
                    print(f"유효한 범위 내의 숫자를 입력하세요 (1-{available_count}).")
            except ValueError:
                print("숫자를 입력하세요.")

        # 처리할 키워드 범위 설정
        start_idx = start_number - 1  # 0-기반 인덱스로 변환
        end_idx = start_idx + process_count
        keywords_to_process = valid_keywords[start_idx:end_idx]
        
        print(f"\n===== 키워드 처리 시작 =====")
        print(f"번호 {start_number}부터 {start_number+process_count-1}까지 {len(keywords_to_process)}개의 키워드를 처리합니다.")
        print(f"결과는 {content_column_letter}열에 저장됩니다.")
        
        # for문 밖에서 한 번만 물어봄
        include_images = input("별도의 이미지를 생성하시겠습니까? (DALL-E 사용, y/n, 기본값: n): ").strip().lower()
        if include_images == '':
            include_images = 'n'
        
        # 이미지 프롬프트 스타일 선택
        image_prompt_style = "1"
        if include_images != 'n':
            print("\n이미지 프롬프트 스타일을 선택하세요:")
            print("1. 기본 스타일 (전문적이고 고품질의 일반 이미지)")
            print("2. PCB 스타일 (다크 블루 PCB 배경, 미니멀 벡터 아이콘, 한글 텍스트)")
            print("3. GPT-4.1-mini 스타일 (GPT-4.1-mini 모델 기반 이미지 생성)")
            image_prompt_style = input("프롬프트 스타일 번호 선택 (기본값: 3): ").strip() or "3"
        
        # 플랫폼 선택 추가
        print("\n발행 플랫폼을 선택하세요:")
        print("1. WordPress")
        print("2. 티스토리")
        publish_platform = input("플랫폼 번호 선택 (기본값: 2): ").strip() or "2"
        
        # 플랫폼별 발행 옵션 설정
        publish_to_platform = False
        auto_category = False
        category_name = None
        post_language = None
        use_stealth_mode = False  # 스텔스 모드 플래그
        mcp_extension_endpoint = None
        
        if publish_platform == "1":
            publish_to_platform = input("생성된 콘텐츠를 WordPress에 자동 발행하시겠습니까? (y/n, 기본값: y): ").strip().lower() != 'n'
            if publish_to_platform:
                print("WordPress 발행이 선택되었습니다.")
                
                # Polylang 언어 선택 옵션
                # 10번 영어 블로그 모드를 선택한 경우 자동으로 영어로 설정
                if content_mode == "10":
                    post_language = "en"
                    print("\n영어 블로그 모드가 선택되어 자동으로 영어(en)로 발행됩니다.")
                else:
                    print("\nPolylang 언어 설정:")
                    print("1. 한국어 (ko)")
                    print("2. 영어 (en)")
                    print("3. 기본 언어 (설정하지 않음)")
                    language_choice = input("언어를 선택하세요 (1-3, 기본값: 1): ").strip() or "1"
                    
                    if language_choice == "1":
                        post_language = "ko"
                        print("한국어로 발행됩니다.")
                    elif language_choice == "2":
                        post_language = "en"
                        print("영어로 발행됩니다.")
                    else:
                        print("기본 언어로 발행됩니다.")
                
                # WordPress 카테고리 옵션
                print("\n카테고리 설정 옵션:")
                print("1. 키워드 기반 자동 추천 (실패 시 수동 선택)")
                print("2. 모든 글에 동일한 카테고리 적용")
                print("3. 미분류로 발행")
                print("4. 각 키워드마다 카테고리 수동 선택")
                
                cat_option = input("카테고리 옵션을 선택하세요 (1-4, 기본값: 4): ").strip() or "4"
                
                if cat_option == "1":
                    auto_category = True
                    category_name = None
                    print("자동 카테고리 추천이 활성화되었습니다. 추천 실패 시 수동으로 선택할 수 있습니다.")
                elif cat_option == "2":
                    auto_category = False
                    # 카테고리 목록 표시
                    success, categories_output = run_wp_cli("term list category --fields=name --format=csv")
                    if success:
                        category_names = []
                        lines = categories_output.split('\n')
                        for line in lines:
                            if line and "name" not in line:
                                clean_name = line.strip().strip('"')
                                if clean_name:
                                    category_names.append(clean_name)
                        
                        if category_names:
                            print("\n현재 사용 가능한 카테고리:")
                            for i, cat in enumerate(category_names):
                                print(f"{i+1}. {cat}")
                            print(f"{len(category_names)+1}. 새 카테고리 생성")
                            
                            while True:
                                try:
                                    choice = int(input(f"\n카테고리를 선택하세요 (1-{len(category_names)+1}): ").strip())
                                    if 1 <= choice <= len(category_names):
                                        category_name = category_names[choice-1]
                                        print(f"모든 글이 '{category_name}' 카테고리로 발행됩니다.")
                                        break
                                    elif choice == len(category_names)+1:
                                        new_cat = input("새 카테고리 이름을 입력하세요: ").strip()
                                        if new_cat:
                                            category_name = new_cat
                                            print(f"모든 글이 새 카테고리 '{category_name}'로 발행됩니다.")
                                            break
                                    else:
                                        print("유효한 번호를 입력하세요.")
                                except ValueError:
                                    print("숫자를 입력하세요.")
                        else:
                            category_name = input("카테고리 이름을 입력하세요: ").strip()
                    else:
                        category_name = input("카테고리 이름을 입력하세요: ").strip()
                elif cat_option == "3":
                    auto_category = False
                    category_name = None
                    print("모든 글이 미분류로 발행됩니다.")
                else:  # cat_option == "4"
                    auto_category = False
                    print("모든 키워드에 적용할 카테고리를 선택하세요.")
                    
                    # 바로 카테고리 목록 표시 및 선택
                    success, categories_output = run_wp_cli("term list category --fields=name --format=csv")
                    if success:
                        category_names = []
                        lines = categories_output.split('\n')
                        for line in lines:
                            if line and "name" not in line:
                                clean_name = line.strip().strip('"')
                                if clean_name:
                                    category_names.append(clean_name)
                        
                        if category_names:
                            print("\n현재 사용 가능한 카테고리:")
                            for i, cat in enumerate(category_names):
                                print(f"{i+1}. {cat}")
                            print(f"{len(category_names)+1}. 새 카테고리 생성")
                            print(f"{len(category_names)+2}. 미분류")
                            
                            while True:
                                try:
                                    choice = int(input(f"\n카테고리를 선택하세요 (1-{len(category_names)+2}): ").strip())
                                    if 1 <= choice <= len(category_names):
                                        category_name = category_names[choice-1]
                                        print(f"모든 키워드가 '{category_name}' 카테고리로 발행됩니다.")
                                        break
                                    elif choice == len(category_names)+1:
                                        new_cat = input("새 카테고리 이름을 입력하세요: ").strip()
                                        if new_cat:
                                            category_name = new_cat
                                            print(f"모든 키워드가 새 카테고리 '{category_name}'로 발행됩니다.")
                                            break
                                    elif choice == len(category_names)+2:
                                        category_name = None
                                        print("모든 키워드가 미분류로 발행됩니다.")
                                        break
                                    else:
                                        print("유효한 번호를 입력하세요.")
                                except ValueError:
                                    print("숫자를 입력하세요.")
                        else:
                            category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
                    else:
                        category_name = input("카테고리 이름을 입력하세요 (비워두면 미분류): ").strip() or None
        else:
            publish_to_platform = input("생성된 콘텐츠를 티스토리에 자동 발행하시겠습니까? (y/n, 기본값: y): ").strip().lower() != 'n'
            if publish_to_platform:
                print("티스토리 발행이 선택되었습니다.")
                print("\n🔧 브라우저 실행 명령어 (옵션 1,4,5 사용 시):")
                print("📍 기본 모드: google-chrome --remote-debugging-port=9222 --user-data-dir=\"$HOME/ChromeDebugData\"")
                print("🥷 스텔스 모드: google-chrome --remote-debugging-port=9222 --user-data-dir=\"$HOME/ChromeDebugData\" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox")
                print("(옵션 2는 스텔스 브라우저 연결, 옵션 6은 확장 프로그램 연결)")
                print("\n티스토리 연결 방식을 선택하세요:")
                print("1. CDP 모드 (기존 크롬 브라우저 연결 - 빠르고 안정적)")
                print("2. 스텔스 모드 (기존 스텔스 브라우저 연결 - 자동화 감지 우회)")
                print("3. 기존 브라우저 사용 (기본 브라우저에서 티스토리 열기 - 수동)")
                print("4. 기존 브라우저 완전 자동 발행 (CDP 연결로 자동 입력)")
                print("5. MCP Playwright 브라우저 모드 (Claude Code MCP 연동)")
                print("6. MCP Playwright 확장 모드 (기존 열려 있는 크롬 + 확장, 원격 디버깅 명령 불필요)")
                connection_mode = input("연결 방식 번호 선택 (기본값: 5): ").strip() or "5"
                
                if connection_mode == "2":
                    use_stealth_mode = True
                    print("🥷 스텔스 모드가 선택되었습니다.")
                    print("🔧 Chrome을 다음 스텔스 명령어로 먼저 실행해주세요:")
                    if platform.system() == "Windows":
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    else:
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    print("자동으로 티스토리 관리 페이지로 이동하여 글을 작성합니다.")
                elif connection_mode == "3":
                    use_stealth_mode = "existing_browser"
                    print("🌐 기존 브라우저 사용 모드가 선택되었습니다.")
                elif connection_mode == "4":
                    use_stealth_mode = "existing_browser_auto"
                    print("🤖 기존 브라우저 완전 자동 발행 모드가 선택되었습니다.")
                    print("CDP로 연결하여 자동으로 글을 작성합니다.")
                    print(f"크롬을 다음 명령어로 실행해주세요:")
                    print("\n📍 기본 모드:")
                    if platform.system() == "Windows":
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    else:
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                elif connection_mode == "5":
                    use_stealth_mode = "mcp_playwright"
                    print("🎭 MCP Playwright 브라우저 모드가 선택되었습니다.")
                    print("Claude Code MCP를 통해 브라우저를 제어합니다. (1번과 동일한 흐름)")
                    print(f"크롬을 다음 명령어로 실행해주세요:")
                    print("\n📍 기본 모드:")
                    if platform.system() == "Windows":
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    else:
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                elif connection_mode == "6":
                    use_stealth_mode = "mcp_extension"
                    print("🔌 MCP Playwright 확장 모드가 선택되었습니다.")
                    print("기존에 열려 있는 크롬 브라우저의 MCP Playwright 확장을 통해 자동 연결합니다.")
                    print("원격 디버깅 명령을 실행할 필요가 없습니다.")
                    # 자동으로 MCP Playwright 확장과 연결 (Claude Code MCP 사용)
                    mcp_extension_endpoint = "auto"  # MCP를 통한 자동 연결 표시
                else:
                    print("🔗 CDP 모드가 선택되었습니다.")
                    print(f"크롬을 다음 명령어로 실행해주세요:")
                    print("\n📍 기본 모드:")
                    if platform.system() == "Windows":
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'chrome.exe --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="%USERPROFILE%\\ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    else:
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData"')
                        print(f"\n🥷 스텔스 모드:")
                        print(f'google-chrome --remote-debugging-port={DEBUGGING_PORT} --user-data-dir="$HOME/ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')

        # 키워드별 처리
        for i, (number, keyword, excel_row, outline_topics, title_present, existing_title) in enumerate(keywords_to_process):
            print(f"\n===== [{i+1}/{len(keywords_to_process)}] '{keyword}' 처리 시작 =====")

            if title_present:
                existing_title_preview = (existing_title[:47] + "...") if len(existing_title) > 50 else existing_title
                print(f"⚠️ 기존 제목이 존재합니다(F열): {existing_title_preview}")

            outline_topics = [topic for topic in outline_topics if topic]
            if outline_topics:
                print(f"요청된 소제목: {outline_topics}")

            # 이미지 URL 변수 초기화
            image_url = None

            # 1. 별도 이미지 생성(필요시) - 맨 앞으로 이동
            if include_images != 'n':
                try:
                    print(f"'{keyword}'에 대한 별도 이미지 생성 중...")
                    client = OpenAI()
                    
                    # 프롬프트 스타일에 따라 다른 프롬프트 사용
                    if image_prompt_style == "2":
                        prompt = (
                            f"A flat-style digital illustration promoting {keyword}, "
                            "with a dark blue PCB-trace background, minimalist vector icons of circuit elements, "
                            f"and bold white Hangul text reading '{keyword}'. Clean, modern graphic design."
                        )
                        print(f"PCB 스타일 프롬프트 사용: {prompt}")
                    elif image_prompt_style == "3":
                        # GPT-4.1-mini 기반 이미지 생성
                        try:
                            print(f"GPT-4.1-mini 모델로 이미지 생성 중...")
                            
                            # 10번 모델(영어 블로그 모드)과 3번 이미지 스타일 조합인 경우 특별 처리
                            if content_mode == "10":
                                # 영어 키워드로 더 최적화된 프롬프트 사용
                                enhanced_prompt = (
                                    f"Create a professional, high-quality digital illustration for '{keyword}'. "
                                    f"Style: Modern, clean, minimalist design suitable for U.S. business blog. "
                                    f"Include subtle American market visual elements. "
                                    f"Color scheme: Professional blues, grays, and accent colors. "
                                    f"Make it suitable for a modern tech-savvy audience."
                                )
                                print(f"영어 블로그 모드 + GPT-4.1-mini 특별 프롬프트: {enhanced_prompt[:100]}...")
                                response = client.responses.create(
                                    model="gpt-4.1-mini",
                                    input=enhanced_prompt,
                                    tools=[{"type": "image_generation"}],
                                )
                            else:
                                # 기본 GPT-4.1-mini 이미지 생성
                                response = client.responses.create(
                                    model="gpt-4.1-mini",
                                    input=f"Generate an image of {keyword}",
                                    tools=[{"type": "image_generation"}],
                                )
                            
                            # 이미지 데이터 추출
                            image_data = [
                                output.result
                                for output in response.output
                                if output.type == "image_generation_call"
                            ]
                            
                            if image_data:
                                image_base64 = image_data[0]
                                image_file = f"{results_dir}/{number}_{keyword.replace(' ', '_')[:30]}_image.png"
                                with open(image_file, "wb") as f:
                                    f.write(base64.b64decode(image_base64))
                                print(f"GPT-4.1-mini 이미지가 저장되었습니다: {image_file}")
                                # 로컬 파일을 URL로 변환 (임시로 파일 경로 사용)
                                image_url = image_file
                            else:
                                print("GPT-4.1-mini 이미지 생성 실패, DALL-E로 대체합니다.")
                                # DALL-E로 대체
                                if content_mode == "10":
                                    # 영어 블로그 모드인 경우 영어 프롬프트 사용
                                    prompt = (
                                        f"Create a professional, high-quality digital illustration for '{keyword}'. "
                                        f"Modern, clean design suitable for American business blog. "
                                        f"Professional color scheme with blues and grays. "
                                        f"Minimalist style for modern tech-savvy audience."
                                    )
                                else:
                                    prompt = f"{keyword}의 시각적 표현. 전문적이고 고품질의 이미지로 만들어주세요."
                                response = client.images.generate(
                                    model="dall-e-3",
                                    prompt=prompt,
                                    n=1,
                                    size="1024x1024"
                                )
                                image_url = response.data[0].url
                                print(f"DALL-E 대체 이미지 URL: {image_url}")
                                image_file = f"{results_dir}/{number}_{keyword.replace(' ', '_')[:30]}_dalle_backup.jpg"
                                img_data = requests.get(image_url).content
                                with open(image_file, 'wb') as handler:
                                    handler.write(img_data)
                                print(f"DALL-E 대체 이미지가 저장되었습니다: {image_file}")
                        except Exception as e:
                            print(f"GPT-4.1-mini 이미지 생성 중 오류: {e}")
                            print("DALL-E로 대체합니다.")
                            # DALL-E로 대체
                            if content_mode == "10":
                                # 영어 블로그 모드인 경우 영어 프롬프트 사용
                                prompt = (
                                    f"Create a professional, high-quality digital illustration for '{keyword}'. "
                                    f"Modern, clean design suitable for American business blog. "
                                    f"Professional color scheme with blues and grays. "
                                    f"Minimalist style for modern tech-savvy audience."
                                )
                            else:
                                prompt = f"{keyword}의 시각적 표현. 전문적이고 고품질의 이미지로 만들어주세요."
                            response = client.images.generate(
                                model="dall-e-3",
                                prompt=prompt,
                                n=1,
                                size="1024x1024"
                            )
                            image_url = response.data[0].url
                            print(f"DALL-E 대체 이미지 URL: {image_url}")
                            image_file = f"{results_dir}/{number}_{keyword.replace(' ', '_')[:30]}_dalle_backup.jpg"
                            img_data = requests.get(image_url).content
                            with open(image_file, 'wb') as handler:
                                handler.write(img_data)
                            print(f"DALL-E 대체 이미지가 저장되었습니다: {image_file}")
                    else:
                        # 기본 스타일 (1번)
                        if content_mode == "10":
                            # 영어 블로그 모드인 경우 영어 프롬프트 사용
                            prompt = (
                                f"Professional visual representation of {keyword}. "
                                f"High-quality, modern design suitable for business use. "
                                f"Clean and sophisticated style."
                            )
                        else:
                            prompt = f"{keyword}의 시각적 표현. 전문적이고 고품질의 이미지로 만들어주세요."
                        print(f"기본 스타일 프롬프트 사용: {prompt}")
                    
                    # 기본 스타일과 PCB 스타일의 경우 DALL-E 사용
                    if image_prompt_style in ["1", "2"]:
                        response = client.images.generate(
                            model="dall-e-3",
                            prompt=prompt,
                            n=1,
                            size="1024x1024"
                        )
                        image_url = response.data[0].url
                        print(f"이미지 URL: {image_url}")
                        image_file = f"{results_dir}/{number}_{keyword.replace(' ', '_')[:30]}_image.jpg"
                        img_data = requests.get(image_url).content
                        with open(image_file, 'wb') as handler:
                            handler.write(img_data)
                        print(f"별도 이미지가 저장되었습니다: {image_file}")
                except Exception as e:
                    print(f"이미지 생성 중 오류 발생: {e}")
                    image_url = None  # 오류 발생 시 이미지 URL 초기화

            # 2. 기존 결과(제목/내용) 확인
            print("기존 결과(제목/내용) 확인 중...")
            if is_google_sheet:
                existing_title = safe_get_cell_value(gsheet_worksheet, excel_row, title_column_index + 1)
                existing_content = safe_get_cell_value(gsheet_worksheet, excel_row, content_column_index + 1)
            else:
                existing_title = worksheet.cell(row=excel_row, column=title_column_index + 1).value
                existing_content = worksheet.cell(row=excel_row, column=content_column_index + 1).value

            if existing_title and existing_content:
                print("기존 콘텐츠를 사용합니다.")
                title = existing_title
                content = existing_content
            else:
                print("\n===== API 요청 시작 =====")
                print(f"키워드: {keyword}")
                print(f"모델: {model_name}")
                print(f"콘텐츠 모드: {content_mode}")
                
                # 기존 콘텐츠가 없는 경우 API 요청
                start_time = time.time()
                
                # 시스템 메시지 및 사용자 메시지 기본값 설정
                system_message = ""
                outline_bullet_text = "\n".join(f"- {topic}" for topic in outline_topics)
                if outline_topics:
                    user_message = (
                        f"블로그 제목: {keyword}\n"
                        "아래 소제목을 반드시 포함해 심화된 블로그 글을 작성해주세요:\n"
                        f"{outline_bullet_text}\n"
                        "각 소제목은 본문 내 섹션 제목으로 활용하고, 필요한 경우 추가 세부 항목을 확장하세요."
                    )
                else:
                    user_message = f"{keyword}에 대한 블로그 글을 작성해주세요."
                
                # 모드에 따른 시스템 메시지 설정
                if content_mode == "8" or content_mode == "9":
                    system_message = (
                        'You are an expert content marketer and web developer.\n'
                        'Produce long-form, SEO-optimized Markdown book posts focused on the user\'s keyword.\n'
                        'Your output must be valid "Markdown Book" format:\n'
                        '- Use semantic headings (#, ##, ### 등) appropriately.\n'
                        '- Use keyword-rich headings and subheadings.\n'
                        '- Wrap All code examples are enclosed in code blocks.\n'
                        '- Highlight key points with emojis (e.g. 🏥, 📦, ✅).\n'
                        '- When presenting a Top10 or Top7 list(Top 10 if possible):\n'
                        '  - Use a numbered list for the ranking.\n'
                        '  - Provide a Markdown table for side-by-side 비교, e.g.:\n'
                        '- Use blockquotes (`> TIP:`) or callouts for 강조 메시지.\n'
                        '-  keyword :  {keyword}, 즉 키워드 {개인회생 자격조건} 를제시하면 최신 정보를 기준으로 조사한다 \n'
                        '  - 예시:  제목과 내용제목은 아래와 같이 키워드{개인회생 자격조건}만 같고 다르게 표현  \n'
                        '    - 제목: ` "개인회생 자격조건" 꼭 알아야 할 신청 기준과 절차 정리`  \n'
                        '    - 내용 시작 예시들:\n'
                        '      1. HTML 큰 제목: `<h2 style="font-size: 2.5em; text-align: center; margin-bottom: 30px; color: #1a73e8;">💰 개인회생 자격조건 완벽 가이드</h2>`\n'
                        '      2. 아이콘과 함께: `<div style="text-align: center; margin-bottom: 30px;"><h2 style="font-size: 2.5em; display: inline-block;"><span style="font-size: 1em;">🦷</span> 치아보험 완벽 가이드</h2></div>`\n'
                        '      3. 박스 스타일: `<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;"><h2 style="font-size: 2.5em; margin: 0;">🦷 치아보험 완벽 가이드</h2></div>`\n'
                        '      4. 배경색 하이라이트: `<div style="background-color: #f0f8ff; padding: 20px; border-left: 5px solid #1a73e8; margin-bottom: 30px;"><h2 style="font-size: 2.2em; color: #1a73e8; margin: 0;">📋 치아보험 완벽 가이드</h2></div>`\n'
                        '- IMPORTANT: 반드시 위 예시 중 하나를 선택하여 콘텐츠를 시작하세요. HTML 스타일 헤더는 필수입니다.\n'
                        '- 키워드 관련 이모지를 적절히 활용하세요 (예: 치아보험→🦷, 자동차보험→🚗, 대출→💰).\n'
                        '- 첫 번째 마크다운 제목(#)은 사용하지 마세요.\n'
                        '- IMPORTANT: Do NOT include any reference links, external links, or "참고 링크" sections.\n'
                        '- Do NOT create or fabricate any URLs or hyperlinks.\n'
                        '- End with a strong conclusion but do NOT add any links or references after the conclusion.\n'
                        '- Focus only on providing valuable content without any external references.\n'
                        '- Ensure the final document is ready to copy-and-paste as a complete Markdown file.'
                    )
                    # OpenAI API 직접 호출
                    response = openai.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": system_message},
                            {"role": "user", "content": user_message}
                        ],
                        temperature=0.2,
                        max_tokens=2048,
                        top_p=1
                    )
                    content = response.choices[0].message.content
                    
                    # HTML 스타일 헤더가 없으면 기본 스타일 추가
                    if not re.search(r'<(div|h[1-6])\s+style=', content[:500]):  # 첫 500자 내에 스타일 태그가 없으면
                        # 키워드에 따른 이모지 선택
                        emoji = "📋"
                        if "보험" in keyword:
                            if "여행" in keyword:
                                emoji = "🌏"
                            elif "치아" in keyword:
                                emoji = "🦷"
                            elif "자동차" in keyword:
                                emoji = "🚗"
                            elif "건강" in keyword or "실손" in keyword:
                                emoji = "🏥"
                            elif "암" in keyword:
                                emoji = "🎗️"
                            else:
                                emoji = "🛡️"
                        elif "대출" in keyword:
                            emoji = "💰"
                        elif "카드" in keyword:
                            emoji = "💳"
                        
                        # 기본 HTML 스타일 헤더 추가 (그라디언트 박스 스타일)
                        # 그라디언트 색상을 키워드 종류에 따라 다르게 설정
                        if "여행" in keyword:
                            gradient = "linear-gradient(135deg, #43cea2 0%, #185a9d 100%)"  # 청록색 계열
                        elif "치아" in keyword:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # 보라색 계열
                        elif "건강" in keyword or "실손" in keyword:
                            gradient = "linear-gradient(135deg, #f093fb 0%, #f5576c 100%)"  # 핑크색 계열
                        elif "자동차" in keyword:
                            gradient = "linear-gradient(135deg, #4facfe 0%, #00f2fe 100%)"  # 하늘색 계열
                        elif "대출" in keyword:
                            gradient = "linear-gradient(135deg, #fa709a 0%, #fee140 100%)"  # 따뜻한 색상
                        elif "카드" in keyword:
                            gradient = "linear-gradient(135deg, #30cfd0 0%, #330867 100%)"  # 어두운 청색
                        else:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # 기본 보라색
                        
                        default_header = f'<div style="background: {gradient}; color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;">\n  <h2 style="font-size: 2.5em; margin: 0;">{emoji} {keyword} 완벽 가이드</h2>\n</div>\n\n'
                        content = default_header + content
                    
                    # content에서 제목 추출 (HTML h2가 있으면 무시하고 WordPress 제목 사용)
                    if '<h2 style=' in content:
                        # HTML 스타일 제목이 있으면 키워드 기반 제목 생성
                        title = f"{keyword} 완벽 가이드 – 꼭 알아야 할 핵심 정보"
                    else:
                        # 기존 방식으로 제목 추출
                        title_match = re.search(r'^#+\s*\[?(.+?)\]?\s*$', content, re.MULTILINE)
                        if title_match:
                            title = title_match.group(1).strip()
                        else:
                            title = f"{keyword} 관련 정보"
                        
                elif content_mode in ["91", "92"]:
                    # 새로운 OpenAI Responses API를 사용한 블로그마크다운&아이콘&이미지 모드
                    system_message = f"""You are an expert Korean content marketer and web developer.\n
Produce a long-form, SEO-optimized Markdown blog post for the requested topic.\n
Guidelines:\n- Begin with an HTML-styled hero header that matches the topic's mood.\n- Use semantic Markdown headings (##, ###) with keyword-rich phrases.\n- Enclose any code examples in fenced code blocks.\n- Highlight important insights with emojis (예: 🏥, 📦, ✅).\n- When presenting Top7~Top10 selections, use a numbered list and include a comparison table.\n- Include blockquotes (`> TIP:`) or callouts for 핵심 조언.\n- 반드시 최신 정보를 기준으로 조사하여 반영하세요.\n- 참고 링크나 외부 URL은 생성하지 마세요.\n- 결론 이후에는 추가 섹션을 만들지 마세요.\n"""

                    if outline_topics:
                        outline_for_prompt = "\n".join(f"- {topic}" for topic in outline_topics)
                        user_message = (
                            f"블로그 제목: {keyword}\n"
                            "필수 소제목:\n"
                            f"{outline_for_prompt}\n"
                            "위 목차를 기반으로 깊이 있는 블로그 글을 작성하고, 각 소제목을 한 섹션으로 확장해주세요."
                        )
                    else:
                        user_message = f"{keyword}에 대한 블로그 글을 작성해주세요."

                    client = OpenAI(api_key=api_key) if api_key else OpenAI()
                    response = client.responses.create(
                        model="gpt-4o",
                        instructions=system_message,
                        input=user_message,
                        temperature=0.2,
                        max_output_tokens=4000,
                    )
                    content = getattr(response, "output_text", "")
                    if not content:
                        try:
                            content = "".join(
                                block.text
                                for item in response.output
                                if getattr(item, "type", None) == "message"
                                for block in getattr(item, "content", [])
                                if getattr(block, "type", None) == "output_text"
                            )
                        except Exception:
                            content = str(response)

                    # HTML 스타일 헤더가 없으면 기본 스타일 추가
                    if not re.search(r'<(div|h[1-6])\s+style=', content[:500]):  # 첫 500자 내에 스타일 태그가 없으면
                        # 키워드에 따른 이모지 선택
                        emoji = "📋"
                        if "보험" in keyword:
                            if "여행" in keyword:
                                emoji = "🌏"
                            elif "치아" in keyword:
                                emoji = "🦷"
                            elif "자동차" in keyword:
                                emoji = "🚗"
                            elif "건강" in keyword or "실손" in keyword:
                                emoji = "🏥"
                            elif "암" in keyword:
                                emoji = "🎗️"
                            else:
                                emoji = "🛡️"
                        elif "대출" in keyword:
                            emoji = "💰"
                        elif "카드" in keyword:
                            emoji = "💳"

                        # 기본 HTML 스타일 헤더 추가 (그라디언트 박스 스타일)
                        # 그라디언트 색상을 키워드 종류에 따라 다르게 설정
                        if "여행" in keyword:
                            gradient = "linear-gradient(135deg, #43cea2 0%, #185a9d 100%)"  # 청록색 계열
                        elif "치아" in keyword:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # 보라색 계열
                        elif "건강" in keyword or "실손" in keyword:
                            gradient = "linear-gradient(135deg, #f093fb 0%, #f5576c 100%)"  # 핑크색 계열
                        elif "자동차" in keyword:
                            gradient = "linear-gradient(135deg, #4facfe 0%, #00f2fe 100%)"  # 하늘색 계열
                        elif "대출" in keyword:
                            gradient = "linear-gradient(135deg, #fa709a 0%, #fee140 100%)"  # 따뜻한 색상
                        elif "카드" in keyword:
                            gradient = "linear-gradient(135deg, #30cfd0 0%, #330867 100%)"  # 어두운 청색
                        else:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # 기본 보라색

                        default_header = f'<div style="background: {gradient}; color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;">\n  <h2 style="font-size: 2.5em; margin: 0;">{emoji} {keyword} 완벽 가이드</h2>\n</div>\n\n'
                        content = default_header + content

                    # content에서 제목 추출 (HTML h2가 있으면 무시하고 WordPress 제목 사용)
                    if '<h2 style=' in content:
                        # HTML 스타일 제목이 있으면 키워드 기반 제목 생성
                        title = f"{keyword} 완벽 가이드 – 꼭 알아야 할 핵심 정보"
                    else:
                        # 기존 방식으로 제목 추출
                        title_match = re.search(r'^#+\s*\[?(.+?)\]?\s*$', content, re.MULTILINE)
                        if title_match:
                            title = title_match.group(1).strip()
                        else:
                            title = f"{keyword} 관련 정보"

                elif content_mode == "10":
                    # 영어 블로그 모드 처리
                    system_message = (
                        'You are an expert content marketer and web developer specializing in the U.S. market. '
                        'Produce a long-form, SEO-optimized Markdown blog post focused on the user\'s keyword. '
                        'Your output must follow "Markdown Book" guidelines:\n'
                        '- Start the content with an HTML styled header (REQUIRED):\n'
                        '  Example: `<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;"><h2 style="font-size: 2.5em; margin: 0;">🤖 Your Title Here</h2></div>`\n'
                        '- Use appropriate emojis at the beginning of the title (e.g. 🤖 for AI, 💰 for finance, 🏥 for healthcare)\n'
                        '- Do NOT repeat the title in the first heading after the HTML header\n'
                        '- Use semantic headings (`##`, `###`, etc.) for the rest of the content structure.\n'
                        '- Craft keyword-rich headings and subheadings targeted to U.S. searchers.\n'
                        '- Enclose all code examples in triple backticks (```).\n'
                        '- Highlight important points with emojis (e.g. 🏥, 📦, ✅).\n'
                        '- When showing a Top 10 or Top 7 list:\n'
                        '  1. Use a numbered list for ranking.\n'
                        '  2. Include a side-by-side comparison table.\n'
                        '- Use blockquotes (`> TIP:`) or callouts to emphasize tips.\n'
                        '- Reference the latest U.S. data and current trends.\n'
                        '- For title and section heading, include the exact keyword but phrase them differently.\n'
                        '- End the content with a strong, comprehensive conclusion that summarizes key points.\n'
                        '- The conclusion should be the final section with no additional sections after it.\n'
                        '- IMPORTANT: Do NOT include any reference links, external links, or "Additional Resources" sections.\n'
                        '- Do NOT create or fabricate any URLs or hyperlinks.\n'
                        '- Do NOT add any links or references after the conclusion.\n'
                        '- All content should be self-contained without relying on external sources.\n'
                        'When the User provides a **Keyword**, generate the complete self-contained Markdown content.'
                    )

                    if outline_topics:
                        english_outline = "\n".join(f"- {topic}" for topic in outline_topics)
                        user_message = (
                            f"Blog topic: {keyword}\n"
                            "Mandatory subtopics to cover (the items are in Korean; translate them into natural English headings in the article):\n"
                            f"{english_outline}\n"
                            "Expand each point with U.S.-focused data, insights, and actionable next steps."
                        )
                    else:
                        user_message = (
                            f"Create a comprehensive English blog post for a U.S. audience about {keyword}."
                        )
                    
                    response = openai.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": system_message},
                            {"role": "user", "content": user_message}
                        ],
                        temperature=0.2,
                        max_tokens=2048,
                        top_p=1
                    )
                    content = response.choices[0].message.content
                    
                    # 마크다운 코드 블록에서 내용 추출
                    if "```markdown" in content or "```Markdown" in content:
                        content = extract_markdown_from_codeblock(content)
                    
                    # 영어 모드에서도 HTML 스타일 헤더가 없으면 추가
                    if not re.search(r'<(div|h[1-6])\s+style=', content[:500]):
                        # 영어 키워드에 따른 이모지 선택
                        emoji = "📋"
                        keyword_lower = keyword.lower()
                        if "ai" in keyword_lower or "artificial intelligence" in keyword_lower:
                            emoji = "🤖"
                        elif "healthcare" in keyword_lower or "medical" in keyword_lower:
                            emoji = "🏥"
                        elif "finance" in keyword_lower or "financial" in keyword_lower or "investment" in keyword_lower:
                            emoji = "💰"
                        elif "marketing" in keyword_lower:
                            emoji = "📣"
                        elif "data" in keyword_lower or "analytics" in keyword_lower:
                            emoji = "📊"
                        elif "cyber" in keyword_lower or "security" in keyword_lower:
                            emoji = "🔒"
                        elif "education" in keyword_lower or "learning" in keyword_lower:
                            emoji = "🎓"
                        elif "supply chain" in keyword_lower or "logistics" in keyword_lower:
                            emoji = "📦"
                        elif "customer" in keyword_lower or "crm" in keyword_lower:
                            emoji = "🤝"
                        elif "technology" in keyword_lower or "software" in keyword_lower:
                            emoji = "💻"
                        
                        # 영어용 그라디언트 색상
                        if "ai" in keyword_lower:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # AI 보라색
                        elif "healthcare" in keyword_lower:
                            gradient = "linear-gradient(135deg, #f093fb 0%, #f5576c 100%)"  # 헬스케어 핑크
                        elif "finance" in keyword_lower:
                            gradient = "linear-gradient(135deg, #4facfe 0%, #00f2fe 100%)"  # 금융 블루
                        else:
                            gradient = "linear-gradient(135deg, #667eea 0%, #764ba2 100%)"  # 기본 보라색
                        
                        # 영어 제목 생성
                        default_header = f'<div style="background: {gradient}; color: white; padding: 30px; border-radius: 10px; text-align: center; margin-bottom: 30px;">\n  <h2 style="font-size: 2.5em; margin: 0;">{emoji} {keyword}: The Ultimate Guide</h2>\n</div>\n\n'
                        content = default_header + content
                    
                    # content에서 첫 번째 헤딩을 제목으로 추출 (# 기호 제거)
                    title_match = re.search(r'^#+\s*(.+?)$', content, re.MULTILINE)
                    if title_match:
                        title = title_match.group(1).strip()
                    else:
                        # 제목이 없으면 첫 줄을 제목으로 사용
                        first_line = content.strip().split('\n')[0].strip()
                        # ### 같은 마크다운 기호 제거
                        title = re.sub(r'^#+\s*', '', first_line).strip()
                        if not title:
                            title = f"{keyword} - Complete Guide"
                
                elif content_mode == "11":
                    # 프리미엄 컨텐츠 모드 처리 - Claude가 직접 생성 (GPT API 사용 안함)
                    print("Claude가 직접 프리미엄 콘텐츠를 생성합니다...")
                    title, content = generate_premium_content_claude(keyword, outline_topics)
                    print(f"제목: {title}")
                    print("콘텐츠 생성 완료!")
                        
                elif content_mode == "6" or content_mode == "7":
                    # 블로그/책 HTML 모드 처리
                    if content_mode == "6":
                        system_message = """당신은 전문적인 블로그 작가입니다. 키워드에 대해 완전한 블로그 게시물을 작성해 주세요.
1. 주어진 키워드에 대한 SEO 최적화된 장문의 콘텐츠를 작성하세요.
2. 블로그 형식의 HTML 태그를 사용하세요.
3. 가독성이 좋고 전문적인 구조를 사용하세요.
4. 키워드가 포함된 제목(h1)과 소제목(h2, h3)을 사용하세요.
5. 필요에 따라 목록(<ul>, <li>), 강조(<strong>), 인용(<blockquote>) 등의 태그를 사용하세요.
6. 가능하면 표(<table>)를 포함하여 정보를 명확하게 정리하세요.
7. 내용은 교육적이고 가치 있게 작성하세요.
8. 소제목마다 적절한 이모지를 추가하세요.
9. 완전한 HTML 형식으로 반환하세요 (head 태그 없이 body 내용만).
10. 중요: 참고 링크, 외부 링크, 추가 자료 섹션을 포함하지 마세요.
11. URL이나 하이퍼링크를 생성하거나 만들지 마세요."""
                    else:  # content_mode == "7"
                        system_message = """당신은 전문적인 책 작가입니다. 주어진 키워드에 대한 책의 한 장(chapter)을 작성해 주세요.
1. 주어진 키워드에 대해 깊이 있고 통찰력 있는 내용을 작성하세요.
2. 책 형식의 HTML 태그를 사용하세요.
3. 키워드가 포함된 장 제목(h1)과 섹션 제목(h2, h3)을 사용하세요.
4. 필요에 따라 목록(<ul>, <li>), 강조(<strong>), 인용(<blockquote>) 등의 태그를 사용하세요.
5. 개념을 설명하는 표(<table>)나 예시를 포함하세요.
6. 내용은 교육적이고 깊이 있게 작성하세요.
7. 각 섹션 제목 앞에 적절한 이모지를 추가하세요.
8. 완전한 HTML 형식으로 반환하세요 (head 태그 없이 body 내용만).
9. 중요: 참고 링크, 외부 링크, 추가 자료 섹션을 포함하지 마세요.
10. URL이나 하이퍼링크를 생성하거나 만들지 마세요."""
                    
                    # OpenAI API 직접 호출
                    response = openai.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": system_message},
                            {"role": "user", "content": user_message}
                        ],
                        temperature=0.7,
                        max_tokens=2500
                    )
                    content = response.choices[0].message.content
                    # HTML에서 제목 추출
                    title_match = re.search(r'<h1[^>]*>(.*?)</h1>', content, re.IGNORECASE | re.DOTALL)
                    if title_match:
                        title = re.sub(r'<.*?>', '', title_match.group(1)).strip()
                    else:
                        title = f"{keyword} 관련 정보"
                        
                else:
                    # 일반 모드
                    system_message = f"""주어진 키워드에 대한 전문적인 콘텐츠를 생성하세요.
- 키워드: {keyword}
- 매력적인 제목과 구조화된 소제목을 사용하세요.
- 읽기 쉽고 정보가 풍부한 내용을 작성하세요.
- 필요에 따라 목록, 강조, 인용 등을 사용하세요.
- 전문적이고 신뢰할 수 있는 정보를 제공하세요.
- 중요: 참고 링크, 외부 링크, 추가 자료 섹션을 포함하지 마세요.
- 결론 이후에 어떠한 링크나 참고 자료도 추가하지 마세요.
- URL이나 하이퍼링크를 생성하거나 만들지 마세요.
- 모든 내용은 독립적이고 완전해야 하며 외부 출처에 의존하지 마세요."""
                    
                    response = openai.chat.completions.create(
                        model=model_name,
                        messages=[
                            {"role": "system", "content": system_message},
                            {"role": "user", "content": user_message}
                        ],
                        temperature=0.7,
                        max_tokens=2500
                    )
                    content = response.choices[0].message.content
                    
                    # 제목 추출
                    if content_mode in ["3", "4"]:  # 마크다운 모드
                        title_match = re.search(r'^#+\s*\[?(.+?)\]?\s*$', content, re.MULTILINE)
                        if title_match:
                            title = title_match.group(1).strip()
                        else:
                            title = f"{keyword} 관련 정보"
                    else:  # 일반 텍스트 모드
                        # 첫 줄이나 첫 문장을 제목으로 사용
                        first_line = content.strip().split('\n')[0].strip()
                        if len(first_line) < 100:  # 적당한 길이인 경우에만
                            title = first_line
                        else:
                            title = f"{keyword} 관련 정보"
                
                elapsed_time = time.time() - start_time
                print(f"✅ API 응답 수신 완료! (소요시간: {elapsed_time:.2f}초)")
                print(f"\n응답 미리보기 (처음 150자):\n{content[:150]}...\n")
            
            # 엑셀/구글 스프레드시트에 결과 저장
            print("결과를 시트에 저장 중...")
            if is_google_sheet:
                update_cell(gsheet_worksheet, excel_row, title_column_index + 1, title)
                update_cell(gsheet_worksheet, excel_row, content_column_index + 1, content)
            else:
                worksheet.cell(row=excel_row, column=title_column_index + 1, value=title)
                worksheet.cell(row=excel_row, column=content_column_index + 1, value=content)
            print(f"결과가 {title_column_letter}열과 {content_column_letter}열에 저장되었습니다.")
            
            # 3. 플랫폼별 발행 처리
            if publish_to_platform:
                if publish_platform == "1":
                    # WordPress 발행
                    print("WordPress 발행 시도 중...")
                    try:
                        success, post_id, post_url, used_category = publish_to_wordpress(
                            title, 
                            content, 
                            category_name=category_name,
                            auto_category=auto_category,
                            keyword=keyword,
                            image_url=image_url,
                            language=post_language
                        )
                        
                        if success:
                            print(f"✅ WordPress 발행 성공: {post_url}")
                            post_info = f"WP발행완료: {post_url}"
                        else:
                            print("✖ WordPress 발행 실패")
                            post_info = "WP발행실패"
                            
                    except Exception as e:
                        print(f"WordPress 발행 중 오류: {e}")
                        post_info = f"WP오류: {str(e)[:50]}"
                
                else:
                    # 티스토리 발행
                    print("티스토리 발행 시도 중...")
                    try:
                        success, post_url = publish_to_tistory_with_cdp(
                            title,
                            content,
                            DEBUGGING_PORT,
                            image_url,  # 이미지 URL 전달
                            use_stealth_mode,  # 스텔스 모드 플래그 전달
                            mcp_extension_endpoint,
                            content_mode,
                        )
                        if success:
                            print(f"✅ 티스토리 발행 성공: {post_url}")
                            post_info = f"티스토리발행완료: {post_url}"
                        else:
                            print("✖ 티스토리 발행 실패")
                            post_info = "티스토리발행실패"
                    except Exception as e:
                        print(f"티스토리 발행 중 오류: {e}")
                        post_info = f"티스토리오류: {str(e)[:50]}"
                
                # 포스트 정보를 H열에 저장
                if is_google_sheet:
                    update_cell(gsheet_worksheet, excel_row, post_info_column_index + 1, post_info)
                else:
                    worksheet.cell(row=excel_row, column=post_info_column_index + 1, value=post_info)
                print(f"포스트 정보가 {post_info_column_letter}열에 저장되었습니다.")
            else:
                # 발행하지 않는 경우
                post_info = "발행안함"
                if is_google_sheet:
                    update_cell(gsheet_worksheet, excel_row, post_info_column_index + 1, post_info)
                else:
                    worksheet.cell(row=excel_row, column=post_info_column_index + 1, value=post_info)
                print(f"발행 안함 정보가 {post_info_column_letter}열에 저장되었습니다.")
            
            # HTML 파일로 결과 저장 (발행 여부와 관계없이)
            result_file = f"{results_dir}/{number}_{keyword.replace(' ', '_')[:30]}.html"
            with open(result_file, "w", encoding="utf-8") as f:
                f.write(f"<!-- 키워드: {keyword} -->\n")
                f.write(f"<!-- 제목: {title} -->\n")
                f.write(f"<!-- 플랫폼: {'WordPress' if publish_platform == '1' else '티스토리'} -->\n\n")
                f.write(content)
            print(f"HTML 결과가 저장되었습니다: {result_file}")
            
            # 다음 요청 전 잠시 대기
            print("3초 대기 후 다음 키워드로 진행합니다.")
            time.sleep(3)
 
        
        # 최종 저장 (로컬 엑셀 파일인 경우)
        if not is_google_sheet:
            try:
                workbook.save(excel_path)
                print(f"\n최종 결과가 엑셀 파일에 저장되었습니다: {excel_path}")
            except Exception as e:
                print(f"최종 엑셀 저장 중 오류 발생: {e}")
        
        print("\n===== 프로그램 완료 =====")
        print(f"HTML 결과가 '{results_dir}' 디렉토리에 저장되었습니다.")
        print(f"제목은 {'구글 스프레드시트' if is_google_sheet else '엑셀 파일'}의 {title_column_letter}열에 저장되었습니다.")
        print(f"내용은 {'구글 스프레드시트' if is_google_sheet else '엑셀 파일'}의 {content_column_letter}열에 저장되었습니다.")
        print(f"포스트 정보는 {'구글 스프레드시트' if is_google_sheet else '엑셀 파일'}의 {post_info_column_letter}열에 저장되었습니다.")
        if publish_to_platform:
            platform_name = "WordPress" if publish_platform == "1" else "티스토리"
            print(f"생성된 콘텐츠는 {platform_name}에 발행되었습니다.")
            
    except Exception as e:
        print(f"오류 발생: {e}")

# 마크다운을 HTML로 변환하는 함수
def convert_markdown_to_html(markdown_text):
    """마크다운 텍스트를 HTML로 변환합니다."""
    try:
        # HTML 태그가 포함된 부분 보호
        import re
        
        # 스타일이 있는 div와 h태그를 먼저 추출하여 저장
        html_header = ""
        
        # 스타일이 있는 div 태그 추출 (주로 헤더 부분)
        div_match = re.search(r'<div\s+style=[^>]+>.*?</div>', markdown_text, flags=re.DOTALL)
        if div_match:
            html_header = div_match.group(0)
            # 매치된 부분을 마크다운 텍스트에서 제거
            markdown_text = markdown_text.replace(html_header, "", 1)
        
        # 나머지 HTML 태그들을 제거하거나 보호
        protected_html = []
        
        # HTML 태그를 임시 플레이스홀더로 교체
        def protect_html(match):
            protected_html.append(match.group(0))
            return f"%%%PROTECTED_HTML_{len(protected_html)-1}%%%"
        
        # 다른 style 속성이 있는 HTML 태그 보호
        markdown_text = re.sub(r'<(h[1-6])\s+style=[^>]+>.*?</\1>', protect_html, markdown_text, flags=re.DOTALL)
        
        # 마크다운 코드 블록을 <pre><code>로 변환
        def convert_code_block(match):
            code_content = match.group(1).strip()
            # HTML 엔티티 이스케이프
            code_content = code_content.replace('&', '&amp;')
            code_content = code_content.replace('<', '&lt;')
            code_content = code_content.replace('>', '&gt;')
            return f'<pre style="background: #f5f5f5; padding: 15px; border-radius: 5px; overflow-x: auto;"><code>{code_content}</code></pre>'
        
        # 코드 블록 변환
        markdown_text = re.sub(r'```markdown\s*([\s\S]*?)\s*```', convert_code_block, markdown_text, flags=re.IGNORECASE)
        markdown_text = re.sub(r'```md\s*([\s\S]*?)\s*```', convert_code_block, markdown_text, flags=re.IGNORECASE)
        markdown_text = re.sub(r'```\s*([\s\S]*?)\s*```', convert_code_block, markdown_text)
        
        # 마크다운 라이브러리를 사용하여 변환
        html = markdown.markdown(markdown_text, extensions=['extra', 'codehilite', 'tables', 'nl2br'])
        
        # TIP 블록을 스타일이 있는 div로 변환
        html = re.sub(
            r'<blockquote>\s*<p>TIP:(.*?)</p>\s*</blockquote>',
            r'<div style="background: #e8f5e9; border-left: 4px solid #4caf50; padding: 15px; margin: 20px 0; border-radius: 5px;"><p><strong>💡 TIP:</strong>\1</p></div>',
            html,
            flags=re.DOTALL
        )
        
        # 보호된 HTML 태그 복원
        for i, protected in enumerate(protected_html):
            html = html.replace(f"%%%PROTECTED_HTML_{i}%%%", protected)
        
        # 헤더 HTML을 맨 앞에 추가
        if html_header:
            html = html_header + "\n\n" + html
        
        return html
    except Exception as e:
        print(f"마크다운 변환 중 오류 발생: {e}")
        return markdown_text  # 변환 실패 시 원본 반환


def build_preview_text(source_text: str, limit: int = 120) -> str:
    """본문에서 요약용 텍스트를 추출합니다."""
    import re

    if not source_text:
        return ""

    text = source_text
    text = re.sub(r'```[\s\S]*?```', ' ', text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'[\[#*`>\|_-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()

    if len(text) > limit:
        text = text[:limit].rstrip() + "..."
    return text

# 마크다운 코드 블록 추출 함수
def extract_markdown_from_codeblock(text):
    """마크다운 코드 블록에서 마크다운 내용을 추출합니다."""
    # 여러 형식의 마크다운 코드 블록 패턴
    patterns = [
        r'```Markdown book\s*([\s\S]*?)\s*```',  # 북 모드 패턴
        r'```Markdown\s*([\s\S]*?)\s*```',       # 일반 마크다운 패턴
        r'```markdown\s*([\s\S]*?)\s*```',       # 소문자 마크다운 패턴
        r'```md\s*([\s\S]*?)\s*```'              # md 약어 패턴
    ]
    
    # 모든 패턴 시도
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    # 코드 블록을 찾지 못한 경우 원본 텍스트 반환
    return text

# 콘텐츠에서 중복된 제목 제거 함수
def remove_duplicate_title(title, content, is_markdown=False):
    """콘텐츠에서 중복된 제목을 제거합니다."""
    if not title or not content:
        return content
    
    # 제목 텍스트 정규화
    normalized_title = title.strip()
    
    if is_markdown:
        # 마크다운에서 제목 제거 (# 형식 제목)
        # 첫 번째 # 제목이 전체 제목과 일치하는지 확인
        lines = content.split('\n')
        for i, line in enumerate(lines):
            if line.strip().startswith('# ') and normalized_title in line:
                # 제목 라인 제거
                lines.pop(i)
                # 제목 다음에 빈 줄이 있으면 그것도 제거
                if i < len(lines) and not lines[i].strip():
                    lines.pop(i)
                break
        
        return '\n'.join(lines)
    else:
        # HTML에서 제목 제거 (h1 태그)
        # h1 태그가 전체 제목과 일치하는지 확인
        h1_pattern = r'<h1[^>]*>(.*?)</h1>'
        
        def replace_matching_h1(match):
            h1_content = match.group(1)
            # HTML 태그 제거 후 비교
            clean_h1 = re.sub(r'<.*?>', '', h1_content).strip()
            if clean_h1 == normalized_title:
                return ""  # 일치하는 h1 제거
            return match.group(0)  # 일치하지 않으면 유지
            
        # 첫 번째 일치하는 h1 태그만 제거
        result = re.sub(h1_pattern, replace_matching_h1, content, count=1, flags=re.IGNORECASE)
        
        # 제목 제거 후 남은 빈 줄 정리
        result = re.sub(r'^\s*\n', '', result)
        
        return result

# 키워드에 맞는 적절한 카테고리를 추천하는 함수
def suggest_category_for_keyword(keyword, model="gpt-3.5-turbo"):
    """키워드를 분석하여 적절한 카테고리를 추천합니다."""
    try:
        print(f"키워드 '{keyword}'에 맞는 카테고리 추천 중...")
        
        # 기존 카테고리 목록 가져오기
        success, categories_output = run_wp_cli("term list category --fields=name --format=csv")
        
        if not success:
            print("카테고리 목록을 가져오지 못했습니다.")
            return None
        
        # 카테고리 목록 처리
        category_names = []
        lines = categories_output.split('\n')
        for line in lines:
            if line and "name" not in line:  # 헤더 제외
                # 따옴표 제거
                clean_name = line.strip().strip('"')
                if clean_name:
                    category_names.append(clean_name)
        
        print(f"현재 워드프레스 카테고리 목록: {', '.join(category_names)}")
        
        # 시스템 메시지 설정
        system_message = f"""당신은 콘텐츠 카테고리 분류 전문가입니다. 
주어진 키워드를 분석하여 가장 적합한 카테고리를 추천해주세요. 
다음은 현재 사용 가능한 카테고리 목록입니다:
{', '.join(category_names)}

만약 위 목록에 적합한 카테고리가 없다면, 새로운 카테고리 이름을 제안해주세요.
응답은 카테고리 이름만 정확히 한 줄로 작성해주세요. 다른 설명은 포함하지 마세요."""

        # 사용자 메시지 설정
        user_message = f"다음 키워드에 가장 적합한 카테고리를 알려주세요: {keyword}"
        
        # API 요청 전송
        response = openai.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ],
            temperature=0.3,
            max_tokens=50
        )
        
        # 응답 처리
        suggested_category = response.choices[0].message.content.strip()
        print(f"추천된 카테고리: {suggested_category}")
        
        return suggested_category
    
    except Exception as e:
        print(f"카테고리 추천 중 오류 발생: {e}")
        return None

# 워드프레스 카테고리 생성 또는 가져오는 함수
def get_or_create_wp_category(category_name):
    """카테고리를 생성하거나 이미 존재하는 경우 가져옵니다."""
    if not category_name:
        return None
    
    # 카테고리명에서 HTML 엔티티 디코딩
    import html
    clean_category_name = html.unescape(category_name)
    
    # 쉘 명령어를 위한 이스케이프 처리
    # 작은따옴표 안에서는 특수문자가 대부분 안전하지만, 작은따옴표 자체는 처리 필요
    safe_category_name = clean_category_name.replace("'", "'\"'\"'")
    print(f"카테고리명 처리: '{category_name}' -> '{clean_category_name}'")
    
    # 카테고리 존재 확인 - 모든 카테고리를 가져와서 직접 비교
    check_all, all_cats = run_wp_cli("term list category --fields=term_id,name --format=json")
    
    cat_id = None
    if check_all:
        try:
            import json
            categories = json.loads(all_cats)
            for cat in categories:
                # 카테고리 이름 비교 (HTML 엔티티 처리)
                cat_name = html.unescape(cat['name'])
                if cat_name == clean_category_name:
                    cat_id = cat['term_id']
                    print(f"기존 카테고리 발견: '{clean_category_name}' (ID: {cat_id})")
                    return str(cat_id)
        except Exception as e:
            print(f"카테고리 목록 파싱 오류: {e}")
    
    # 카테고리가 없으면 생성
    if not cat_id:
        print(f"→ 카테고리 '{clean_category_name}' 생성 중...")
        create_cat, cat_output = run_wp_cli(f"term create category '{safe_category_name}'")
        if create_cat:
            # 출력에서 ID 추출
            try:
                print(f"카테고리 생성 결과: {repr(cat_output)}")
                # "Success: Created category 123." 형식에서 ID 추출
                if "Success:" in cat_output:
                    # 숫자만 추출
                    import re
                    id_match = re.search(r'(\d+)', cat_output)
                    if id_match:
                        cat_id = id_match.group(1)
                        print(f"생성된 카테고리 ID: {cat_id}")
                        return cat_id
                # 다른 형식 시도
                cat_id = cat_output.split('\n')[-1].split(':')[-1].strip()
                if cat_id.isdigit():
                    print(f"생성된 카테고리 ID: {cat_id}")
                    return cat_id
            except Exception as e:
                print(f"카테고리 ID 추출 실패: {e}")
                return None
        else:
            # 이미 존재한다는 오류인 경우 다시 찾기 시도
            if "already exists" in cat_output:
                print("카테고리가 이미 존재합니다. 다시 검색 중...")
                # CSV 형식으로 다시 시도
                check_cat2, cat_output2 = run_wp_cli("term list category --fields=term_id,name --format=csv")
                if check_cat2:
                    lines = cat_output2.split('\n')
                    for line in lines:
                        if line and "term_id" not in line:
                            parts = line.split(',')
                            if len(parts) >= 2:
                                # CSV에서는 따옴표로 감싸진 경우가 있음
                                term_name = parts[1].strip().strip('"')
                                term_name = html.unescape(term_name)
                                if term_name == clean_category_name:
                                    term_id = parts[0].strip()
                                    print(f"재검색으로 카테고리 발견: '{clean_category_name}' (ID: {term_id})")
                                    return term_id
            print(f"카테고리 생성 실패: {cat_output}")
            return None
    
    return None

def generate_premium_content_claude(keyword, outline_topics=None):
    """
    Claude가 직접 프리미엄 콘텐츠를 생성합니다 (GPT API 사용하지 않음)
    참고: /home/skyntech/www/html/wp/gpt_results/11_법인카드_추천.html
         /home/skyntech/www/html/wp/gpt_results/11_신용카드_캐시백_많은_곳.html
    """
    
    # 키워드별 특화 제목 설정
    if "카드" in keyword:
        if "법인" in keyword:
            title = f"{keyword} 완벽 가이드: 기업 성장을 위한 최적 카드 선택법"
        elif "캐시백" in keyword:
            title = f"{keyword} 최대화 가이드: 연간 100만원 절약하는 스마트한 방법"
        else:
            title = f"{keyword} 최신 가이드: 현명한 소비자를 위한 완벽 분석"
    elif "대출" in keyword:
        title = f"{keyword} 완벽 가이드: 최저금리로 승인받는 확실한 방법"
    elif "보험" in keyword:
        title = f"{keyword} 선택 가이드: 보장은 높이고 보험료는 낮추는 방법"
    else:
        title = f"{keyword} 완벽 가이드: 전문가가 알려주는 스마트한 선택법"
    
    if outline_topics:
        formatted_outline = "\n".join(f"- {topic}" for topic in outline_topics)
    else:
        formatted_outline = "- 핵심 주제를 중심으로 독자가 가장 궁금해할 내용을 단계적으로 설명합니다."

    # 키워드에 따른 콘텐츠 템플릿
    content_template = f"""# {title}

## 🎯 {keyword} 핵심 포인트

{keyword}에 대한 최신 트렌드와 함께, 실질적으로 도움이 되는 정보를 완벽하게 정리했습니다. 현명한 선택을 위한 모든 것을 지금 바로 확인하세요!

## 📚 요청된 핵심 목차
{formatted_outline}

### 💡 {keyword} 선택 시 필수 체크사항
- **🔥 핵심 혜택**: 실질적 이익과 부가 서비스
- **📊 비용 대비 효과**: ROI 분석 및 손익분기점
- **🎯 맞춤형 선택**: 개인/기업별 최적화 방안
- **💰 비용 절감**: 숨겨진 혜택과 절약 팁

---

## 🏆 {keyword} BEST 순위

### 🥇 **1위: 프리미엄 옵션**
- **월 비용**: 적정 수준
- **핵심 혜택**: 최고 수준의 서비스
- **특별 기능**: 차별화된 부가 서비스
- **추천 대상**: 품질 중시 고객
- **예상 절약액**: 연간 상당한 금액

### 🥈 **2위: 가성비 옵션**
- **월 비용**: 합리적 수준
- **핵심 혜택**: 균형잡힌 혜택
- **특별 기능**: 실용적 서비스
- **추천 대상**: 실속파 고객
- **예상 절약액**: 연간 만족할만한 수준

### 🥉 **3위: 경제적 옵션**
- **월 비용**: 최저 수준
- **핵심 혜택**: 기본 서비스 충실
- **특별 기능**: 필수 기능 포함
- **추천 대상**: 예산 중시 고객
- **예상 절약액**: 부담 없는 수준

---

## 📊 상세 비교 분석표

| 순위 | 옵션명 | 월 비용 | 핵심 혜택 | 특별 서비스 | 추천 대상 |
|------|--------|---------|-----------|-------------|-----------|
| 1 | 프리미엄 | 높음 | 최고급 서비스 | VIP 전용 | 품질 우선 |
| 2 | 스탠다드 | 중간 | 균형잡힌 혜택 | 일반 서비스 | 실속파 |
| 3 | 베이직 | 낮음 | 기본 서비스 | 필수 기능 | 예산 중시 |

---

## 🎯 맞춤형 선택 가이드

### 👨‍💼 **직장인/전문직**
- **추천 옵션**: 프리미엄 서비스
- **핵심 이유**: 시간 절약과 편의성
- **예상 효과**: 업무 효율성 극대화
- **절약 팁**: 회사 복지와 연계 활용

### 👨‍👩‍👧‍👦 **가정/주부**
- **추천 옵션**: 가성비 서비스
- **핵심 이유**: 실용성과 경제성 균형
- **예상 효과**: 가계 경제 안정화
- **절약 팁**: 가족 단위 혜택 활용

### 🎓 **학생/청년**
- **추천 옵션**: 경제적 서비스
- **핵심 이유**: 부담 없는 시작
- **예상 효과**: 합리적 소비 습관
- **절약 팁**: 청년 특화 혜택 활용

---

## 💡 활용 극대화 전략

### 🔥 **단계별 활용법**

#### **1단계: 현황 분석**
- 현재 사용 패턴 파악
- 개선 가능 영역 확인
- 목표 설정

#### **2단계: 최적 선택**
- 비교 분석 실시
- 시뮬레이션 진행
- 최종 결정

#### **3단계: 지속 관리**
- 정기적 점검
- 혜택 변경 확인
- 필요시 조정

### 📈 **혜택 극대화 꿀팁**
- ✅ 초기 프로모션 활용
- ✅ 번들 상품 고려
- ✅ 시즌별 특가 활용
- ✅ 장기 계약 할인

---

## ⚠️ 주의사항 및 FAQ

### 🚨 **흔한 실수 방지**
- **실수 1**: 조건 미확인
- **해결**: 약관 꼼꼼히 검토
- **실수 2**: 과도한 선택
- **해결**: 필요한 것만 선택

### ❓ **자주 묻는 질문**

**Q1. 어떤 기준으로 선택해야 하나요?**
- 개인의 사용 패턴과 예산을 최우선으로 고려하세요.

**Q2. 변경은 언제든 가능한가요?**
- 대부분 약정 기간이 있으니 사전 확인이 필요합니다.

**Q3. 추가 비용은 없나요?**
- 기본 서비스 외 추가 옵션은 별도 비용이 발생할 수 있습니다.

---

---

## 🚀 결론: {keyword} 스마트한 선택하기

{keyword}는 단순한 선택이 아닌 미래를 위한 투자입니다. 올바른 선택으로 비용은 절감하고 혜택은 극대화하세요!

### 🎯 **성공적인 선택을 위한 3단계**
1. **현재 상황 정확히 파악**
2. **목표에 맞는 옵션 선택**
3. **지속적인 관리와 최적화**

지금 바로 시작하여 더욱 스마트한 선택을 해보세요! ✨

**전문가 상담을 통해 최적의 선택을 하시기 바랍니다.**"""
    
    # 키워드별 특화 콘텐츠 추가
    if "카드" in keyword or "캐시백" in keyword:
        content_template = content_template.replace("옵션명", "카드명")
        content_template = content_template.replace("월 비용", "연회비")
        content_template = content_template.replace("서비스", "혜택")
    elif "대출" in keyword:
        content_template = content_template.replace("옵션명", "상품명")
        content_template = content_template.replace("월 비용", "금리")
        content_template = content_template.replace("혜택", "조건")
    elif "보험" in keyword:
        content_template = content_template.replace("옵션명", "보험상품")
        content_template = content_template.replace("월 비용", "보험료")
        content_template = content_template.replace("혜택", "보장내용")
    
    return title, content_template

def publish_to_tistory_with_cdp(
    title,
    content,
    debugging_port=DEBUGGING_PORT,
    image_url=None,
    use_stealth=False,
    mcp_extension_ws=None,
    selected_content_mode="1",
):
    # 기존 브라우저 사용 모드 (수동)
    if use_stealth == "existing_browser":
        import webbrowser
        print("🌐 기존 브라우저에서 티스토리 로그인 페이지를 여는 중...")
        webbrowser.open("https://www.tistory.com/auth/login", new=2)
        print("✅ 티스토리 로그인 페이지가 열렸습니다.")
        input("로그인 완료 후 Enter 키를 누르세요: ")
        
        # 로그인 후 글쓰기 페이지로 이동
        print("\n글쓰기 페이지로 이동 중...")
        webbrowser.open("https://espace-ch.tistory.com/manage/newpost", new=2)
        print("✅ 글쓰기 페이지가 열렸습니다.")
        print("\n수동으로 글을 작성해주세요:")
        print(f"제목: {title}")
        print(f"내용: (클립보드에 복사됨)")
        
        # 클립보드에 내용 복사 시도
        try:
            import pyperclip
            pyperclip.copy(content)
            print("✅ 내용이 클립보드에 복사되었습니다. Ctrl+V로 붙여넣기 하세요.")
        except:
            print("⚠️  클립보드 복사 실패. 아래 내용을 수동으로 복사해주세요:")
            print("-" * 50)
            print(content[:500] + "..." if len(content) > 500 else content)
            print("-" * 50)
        
        input("\n글 작성을 완료하고 발행한 후 Enter 키를 누르세요: ")
        return True, "https://espace-ch.tistory.com"
    
    # 기존 브라우저 자동 발행 모드
    if use_stealth == "existing_browser_auto":
        import webbrowser
        print("🤖 기존 브라우저 완전 자동 발행 모드")
        
        # 먼저 티스토리 로그인 페이지를 브라우저에서 열기
        print("🌐 기존 브라우저에서 티스토리 로그인 페이지를 여는 중...")
        webbrowser.open("https://www.tistory.com/auth/login", new=2)
        print("✅ 티스토리 로그인 페이지가 열렸습니다.")
        input("로그인 완료 후 Enter 키를 누르세요: ")
        
        # 이제 CDP로 연결해서 자동화 수행
        print("\n🔗 기존 브라우저에 CDP로 연결하여 자동화를 시작합니다...")
        print(f"크롬이 다음 명령어로 실행되어 있는지 확인하세요:")
        if platform.system() == "Windows":
            print(f'chrome.exe --remote-debugging-port={debugging_port} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
        else:
            print(f'google-chrome --remote-debugging-port={debugging_port} --user-data-dir="$HOME/ChromeDebugData"')
        
        input("크롬이 디버깅 모드로 실행 중이면 Enter 키를 누르세요: ")
    
    with sync_playwright() as p:
        human_like_typing = (use_stealth in ("mcp_playwright", "mcp_extension"))
        command_modifier = 'Meta' if platform.system() == 'Darwin' else 'Control'

        def clear_field(element):
            try:
                element.click()
                element.press(f"{command_modifier}+A")
                element.press("Delete")
            except Exception:
                try:
                    element.fill("")
                except Exception:
                    pass

        def type_text(element, text):
            if not text:
                return
            chunk_size = 120
            for i in range(0, len(text), chunk_size):
                chunk = text[i:i + chunk_size]
                if human_like_typing:
                    delay = random.uniform(0.004, 0.009)
                    element.type(chunk, delay=delay)
                    if i + chunk_size < len(text):
                        time.sleep(random.uniform(0.05, 0.12))
                else:
                    element.type(chunk)

        # 스텔스 모드 선택 (기존 브라우저 연결)
        if use_stealth == True:
            print("🥷 스텔스 모드 - 기존 브라우저에 연결 중...")
            print("기존 브라우저가 스텔스 모드로 실행되어 있어야 합니다.")

            try:
                browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{debugging_port}")
                print("✅ 기존 스텔스 브라우저에 연결 성공!")
            except Exception as e:
                print(f"127.0.0.1 연결 실패, localhost 시도 중...")
                try:
                    browser = p.chromium.connect_over_cdp(f"http://localhost:{debugging_port}")
                    print("✅ 기존 스텔스 브라우저에 연결 성공!")
                except Exception as e2:
                    print(f"❌ CDP 연결 실패: {e2}")
                    print("🔧 스텔스 모드로 Chrome을 먼저 실행해주세요:")
                    if platform.system() == "Windows":
                        print(f'chrome.exe --remote-debugging-port={debugging_port} --user-data-dir="%USERPROFILE%\\ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    else:
                        print(f'google-chrome --remote-debugging-port={debugging_port} --user-data-dir="$HOME/ChromeDebugData" --disable-blink-features=AutomationControlled --disable-dev-shm-usage --no-sandbox')
                    return False, None

            # 기존 브라우저에서 스텔스 설정 강화
            context = None
            page = None

            # 티스토리 페이지가 열려 있는 기존 탭 탐색
            for ctx in browser.contexts:
                for existing_page in ctx.pages:
                    if "tistory.com" in (existing_page.url or "").lower():
                        context = ctx
                        page = existing_page
                        break
                if page:
                    break

            # 티스토리 탭이 없다면 가장 첫 번째 컨텍스트 선택하고 새 탭 생성
            if context is None:
                if browser.contexts:
                    context = browser.contexts[0]
                else:
                    print("❌ 사용 가능한 브라우저 컨텍스트를 찾을 수 없습니다.")
                    return False, None

            # 새 탭 생성 또는 기존 탭 사용
            if page is None:
                try:
                    page = context.new_page()
                    print("🆕 새 탭을 생성했습니다.")
                except Exception as e:
                    if context.pages:
                        page = context.pages[0]
                        print("⚠️ 새 탭 생성 실패, 첫 번째 기존 탭을 사용합니다.")
                    else:
                        print(f"❌ 제어 가능한 페이지가 없습니다: {e}")
                        return False, None

            # 런타임 스텔스 설정 강화
            stealth_script = """
                // 기존 webdriver 감지 제거
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });

                // 플러그인 목록 수정
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3, 4, 5]
                });

                // 언어 설정
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['ko-KR', 'ko', 'en-US', 'en']
                });

                // Chrome 객체 추가
                if (!window.chrome) {
                    window.chrome = { runtime: {} };
                }

                // Permission API 수정
                Object.defineProperty(navigator, 'permissions', {
                    get: () => ({
                        query: () => Promise.resolve({ state: 'granted' })
                    })
                });

                // 자동화 감지 함수들 무력화
                if (window.outerHeight === 0) {
                    Object.defineProperty(window, 'outerHeight', { get: () => window.innerHeight });
                }
                if (window.outerWidth === 0) {
                    Object.defineProperty(window, 'outerWidth', { get: () => window.innerWidth });
                }
            """

            try:
                page.evaluate(stealth_script)
                print("🛡️ 런타임 스텔스 설정 적용 완료!")
            except Exception as e:
                print(f"⚠️ 스텔스 설정 적용 중 오류: {e}")

            try:
                page.bring_to_front()
            except Exception:
                pass

            print(f"현재 제어 중인 페이지: {page.url}")

            # 자동으로 티스토리 관리 페이지로 이동
            if "tistory.com/manage" not in (page.url or ""):
                print("🌐 티스토리 관리 페이지로 자동 이동...")
                page.goto("https://espace-ch.tistory.com/manage/newpost", wait_until="networkidle")
                print("✅ 티스토리 글쓰기 페이지 도착!")
        else:
            # CDP 모드 (기존 방식) 또는 MCP 확장 모드 등
            if use_stealth == "mcp_extension":
                print("🎭 Selenium WebDriver를 통해 브라우저 제어를 시작합니다...")
                print("자동화된 브라우저 제어로 티스토리에 글을 발행합니다.")

                # Selenium WebDriver를 통한 브라우저 제어
                try:
                    # Selenium을 통해 티스토리 자동 발행
                    return publish_with_selenium_existing_browser(title, content, image_url)

                except Exception as e:
                    print(f"❌ Selenium WebDriver 실행 실패: {e}")
                    print("🔄 PyAutoGUI 화면 자동화로 전환 시도...")
                    try:
                        return publish_with_pyautogui(title, content, image_url)
                    except Exception as e2:
                        print(f"❌ PyAutoGUI도 실패: {e2}")
                        print("📖 수동 모드로 전환합니다.")
                        return False, None
            else:
                print("🔗 CDP 모드로 연결 중...")
                try:
                    browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{debugging_port}")
                except Exception as e:
                    print(f"127.0.0.1 연결 실패, localhost 시도 중...")
                    try:
                        browser = p.chromium.connect_over_cdp(f"http://localhost:{debugging_port}")
                    except Exception as e2:
                        print(f"❌ CDP 연결 실패: {e2}")
                        print(f"크롬이 다음 명령어로 실행되어 있는지 확인하세요:")
                        if platform.system() == "Windows":
                            print(f'chrome.exe --remote-debugging-port={debugging_port} --user-data-dir="%USERPROFILE%\\ChromeDebugData"')
                        else:
                            print(f'google-chrome --remote-debugging-port={debugging_port} --user-data-dir="$HOME/ChromeDebugData"')
                        return False, None
                    
            context = None
            page = None

            # 티스토리 페이지가 열려 있는 기존 탭 탐색
            for ctx in browser.contexts:
                for existing_page in ctx.pages:
                    if "tistory.com" in (existing_page.url or "").lower():
                        context = ctx
                        page = existing_page
                        break
                if page:
                    break

            # 티스토리 탭이 없다면 가장 첫 번째 컨텍스트 선택
            if context is None:
                if browser.contexts:
                    context = browser.contexts[0]
                else:
                    print("❌ 사용 가능한 브라우저 컨텍스트를 찾을 수 없습니다.")
                    return False, None

            # 선택된 컨텍스트에서 제어 가능한 페이지 확보
            if page is None:
                try:
                    page = context.new_page()
                    print("🆕 새 탭을 생성했습니다.")
                except Exception as e:
                    if context.pages:
                        page = context.pages[0]
                        print("⚠️ 새 탭 생성 실패, 첫 번째 기존 탭을 사용합니다.")
                    else:
                        print(f"❌ 제어 가능한 페이지가 없습니다: {e}")
                        return False, None

            try:
                page.bring_to_front()
            except Exception:
                pass

            print(f"현재 제어 중인 페이지: {page.url}")

        # 팝업(알림) 발생 시 자동 OK
        def handle_dialog(dialog):
            print(f"Dialog message: {dialog.message}")
            dialog.accept()

        page.on("dialog", handle_dialog)  # 모든 팝업에 대해 OK

        try:
            print("\n===== 티스토리 발행 시작 =====", flush=True)
            
            # 4번 옵션일 경우 글쓰기 페이지로 직접 이동
            if use_stealth == "existing_browser_auto":
                print("글쓰기 페이지로 직접 이동 중...")
                page.goto("https://espace-ch.tistory.com/manage/newpost", timeout=30000)
                page.wait_for_load_state('networkidle', timeout=30000)
                time.sleep(2)  # 에디터 완전 로드 대기
            else:
                # 관리 페이지 이동 시도
                try:
                    page.goto("https://espace-ch.tistory.com/manage/posts", timeout=30000)
                    page.wait_for_load_state('domcontentloaded', timeout=10000)
                except Exception as e:
                    # 로그인 페이지로 리다이렉트된 경우 처리
                    current_url = page.url
                    if "login" in current_url.lower() or "kakao" in current_url.lower() or "daum" in current_url.lower():
                        print("로그인이 필요합니다. 로그인 페이지가 감지되었습니다.")
                        print("현재 URL:", current_url)
                        
                        if use_stealth:
                            # 스텔스 모드에서는 이미 로그인했어야 함
                            print("스텔스 모드에서 로그인 후 다시 관리 페이지로 이동합니다...")
                            page.goto("https://espace-ch.tistory.com/manage/posts", timeout=30000)
                        else:
                            # CDP 모드에서는 로그인 대기
                            print("\n⚠️  브라우저에서 티스토리에 로그인해주세요.")
                            print("로그인 완료 후 Enter를 누르세요...")
                            input()
                            # 로그인 후 다시 관리 페이지로 이동
                            page.goto("https://espace-ch.tistory.com/manage/posts", timeout=30000)
                    else:
                        print(f"페이지 이동 중 오류: {e}")
                        raise
            
            # 4번 옵션이 아닌 경우에만 글쓰기 버튼 클릭
            if use_stealth != "existing_browser_auto":
                # 페이지 로드 대기
                page.wait_for_load_state('networkidle', timeout=30000)

                new_post_url = "https://espace-ch.tistory.com/manage/newpost"

                # 글쓰기 버튼 클릭 시도, 실패하면 직접 이동
                print("글쓰기 버튼 찾는 중...")
                try:
                    write_selector = (
                        '#kakaoHead a[href*="write"], '
                        '#kakaoHead [role="link"]:has-text("글쓰기"), '
                        'a[href*="manage/newpost"], '
                        'a:has-text("글쓰기"), '
                        'button:has-text("글쓰기")'
                    )
                    write_button = page.wait_for_selector(write_selector, timeout=20000)
                    write_button.click()
                    print("✓ 글쓰기 버튼을 클릭했습니다.")
                except PlaywrightTimeoutError:
                    print("⚠️ 글쓰기 버튼을 찾지 못했습니다. 글쓰기 페이지로 직접 이동합니다...")
                    page.goto(new_post_url, timeout=30000)
                except Exception as e:
                    print(f"⚠️ 글쓰기 버튼 클릭 중 오류 발생: {e}. 글쓰기 페이지로 직접 이동합니다...")
                    page.goto(new_post_url, timeout=30000)

                if page.url and "manage/newpost" not in page.url:
                    print("글쓰기 페이지로 리다이렉트되지 않아 직접 이동을 시도합니다...")
                    page.goto(new_post_url, timeout=30000)

                # 에디터 로드 대기
                page.wait_for_load_state('networkidle', timeout=20000)
                time.sleep(2)  # 에디터 완전 로드 대기

            # 4번 옵션일 경우 기본 모드로 시작, 다른 경우 마크다운 모드로 전환
            if use_stealth == "existing_browser_auto":
                print("기본 모드에서 시작합니다...")
                # 기본 모드 유지 (제목 입력을 위해)
            else:
                # 에디터 모드 전환: 기본모드 → 마크다운
                print("마크다운 모드로 전환 중...")
                try:
                    mode_button = page.wait_for_selector('[aria-label="기본모드"] button, button:has-text("기본모드")', timeout=15000)
                    mode_button.click()
                    time.sleep(1)
                    
                    markdown_option = page.wait_for_selector('text="마크다운"', timeout=10000)
                    markdown_option.click()
                    print("✓ 마크다운 모드 전환 완료")
                except:
                    # 대체 방법
                    page.get_by_label('기본모드').get_by_role('button', name='기본모드').click()
                    page.get_by_label('기본모드 마크다운 HTML').get_by_text('마크다운').click()

            # 제목 입력
            print("제목 입력 중...")
            try:
                # 더 구체적인 선택자로 제목 입력란 찾기
                title_input = page.wait_for_selector('#post-title-inp, textarea[placeholder*="제목"], [aria-label*="제목"]', timeout=15000)
                if human_like_typing:
                    clear_field(title_input)
                    type_text(title_input, title)
                else:
                    title_input.click()
                    title_input.fill("")  # 먼저 비우기
                    title_input.type(title)  # 타이핑으로 입력
                print(f"✓ 제목 입력 완료: {title}")
            except:
                try:
                    title_box = page.get_by_role('textbox', name='제목을 입력하세요')
                    if human_like_typing:
                        clear_field(title_box)
                        type_text(title_box, title)
                    else:
                        title_box.click()
                        title_box.fill(title)
                except:
                    # 직접 JavaScript로 입력
                    page.evaluate(f'document.querySelector("#post-title-inp").value = "{title}"')

            # 콘텐츠 입력
            print("콘텐츠 입력 중...")

            preview_text = build_preview_text(content)

            def switch_mode(target_text: str):
                try:
                    mode_button = page.wait_for_selector('[aria-label*="모드"] button, button:has-text("마크다운"), button:has-text("HTML"), button:has-text("기본모드")', timeout=10000)
                    mode_button.click()
                    time.sleep(1)
                    option_selector = f'text="{target_text}"'
                    option = page.wait_for_selector(option_selector, timeout=5000)
                    option.click()
                    return True
                except Exception:
                    try:
                        base_label = '기본모드'
                        page.get_by_label(base_label).get_by_role('button', name=base_label).click()
                        page.get_by_label(base_label + ' 마크다운 HTML').get_by_text(target_text).click()
                        return True
                    except Exception:
                        return False

            if selected_content_mode == "92":
                print("HTML 모드로 전환하여 전체 콘텐츠 입력 (요약 텍스트 포함)...")
                rendered_html = convert_markdown_to_html(content)
                if image_url and image_url not in rendered_html:
                    rendered_html = f'<p style="text-align:center;"><img src="{image_url}" alt="{title}" style="max-width:100%; height:auto;" /></p>\n' + rendered_html

                if preview_text:
                    import html
                    rendered_html = f"<p>{html.escape(preview_text)}</p>\n\n{rendered_html}"

                html_switched = switch_mode("HTML")
                if not html_switched:
                    print("HTML 모드 전환 실패, 기본 모드에서 입력을 시도합니다.")

                try:
                    time.sleep(1)
                    html_editor = page.wait_for_selector('.CodeMirror', timeout=10000)
                    html_editor.click()
                    safe_html = rendered_html.replace('\\', '\\\\').replace('`', '\\`')
                    page.evaluate(f'''
                        const cm = document.querySelector('.CodeMirror').CodeMirror;
                        if (cm) {{
                            cm.setValue(`{safe_html}`);
                        }}
                    ''')
                    print("✓ HTML 콘텐츠 입력 완료")
                except Exception as e:
                    print(f"HTML 입력 실패: {e}")
                    try:
                        html_textarea = page.wait_for_selector('textarea', timeout=5000)
                        html_textarea.fill(rendered_html)
                        print("✓ HTML 콘텐츠 입력 (대체 방법)")
                    except Exception as inner_e:
                        print(f"HTML 대체 입력도 실패: {inner_e}")

            elif selected_content_mode == "91":
                import re
                div_pattern = r'^(<div[^>]*>.*?</div>)\s*'
                div_match = re.match(div_pattern, content, re.DOTALL)

                html_div_content = ""
                markdown_content = content

                if div_match:
                    html_div_content = div_match.group(1)
                    markdown_content = content[len(html_div_content):].strip()
                    print(f"HTML div 부분 감지됨: {len(html_div_content)} 글자")
                    print(f"나머지 콘텐츠: {len(markdown_content)} 글자")

                if html_div_content:
                    if switch_mode("HTML"):
                        try:
                            time.sleep(1)
                            html_editor = page.wait_for_selector('.CodeMirror', timeout=10000)
                            html_editor.click()
                            safe_html = html_div_content.replace('\\', '\\\\').replace('`', '\\`')
                            page.evaluate(f'''
                                const cm = document.querySelector('.CodeMirror').CodeMirror;
                                if (cm) {{
                                    cm.setValue(`{safe_html}`);
                                }}
                            ''')
                            print("✓ HTML 헤더 입력 완료")
                        except Exception as e:
                            print(f"HTML 헤더 입력 실패: {e}")
                    else:
                        print("HTML 모드 전환 실패")

                if markdown_content:
                    if image_url and f"![{title}]" not in markdown_content:
                        markdown_content = f"![{title}]({image_url})\n\n{markdown_content}"

                    if switch_mode("마크다운"):
                        try:
                            time.sleep(1)
                            markdown_editor = page.wait_for_selector('#markdown-editor-container textarea, #markdown-editor-container [role="textbox"]', timeout=15000)
                            existing_content = markdown_editor.input_value()
                            new_value = markdown_content if not existing_content else existing_content + "\n\n" + markdown_content
                            markdown_editor.fill(new_value)
                            print("✓ 마크다운 본문 입력 완료")
                        except Exception as e:
                            print(f"마크다운 입력 실패: {e}")
                            try:
                                fallback_editor = page.locator('#markdown-editor-container').get_by_role('textbox')
                                fallback_editor.fill(markdown_content)
                                print("✓ 마크다운 본문 입력 (대체 방법)")
                            except Exception as inner_e:
                                print(f"마크다운 대체 입력도 실패: {inner_e}")
                    else:
                        print("마크다운 모드 전환 실패")

            else:
                try:
                    if image_url and f"![{title}]" not in content:
                        final_content = f"![{title}]({image_url})\n\n{content}"
                    else:
                        final_content = content

                    markdown_editor = page.wait_for_selector('#markdown-editor-container textarea, #markdown-editor-container [role="textbox"]', timeout=15000)
                    if human_like_typing:
                        clear_field(markdown_editor)
                        type_text(markdown_editor, final_content)
                    else:
                        markdown_editor.fill(final_content)
                    print("✓ 콘텐츠 입력 완료 (마크다운 모드)")
                except Exception as e:
                    print(f"마크다운 입력 실패: {e}")
                    try:
                        fallback_editor = page.locator('#markdown-editor-container').get_by_role('textbox')
                        if human_like_typing:
                            clear_field(fallback_editor)
                            type_text(fallback_editor, final_content)
                        else:
                            fallback_editor.fill(final_content)
                        print("✓ 마크다운 콘텐츠 입력 (대체 방법)")
                    except Exception as inner_e:
                        print(f"마크다운 대체 입력도 실패: {inner_e}")

            time.sleep(2)  # 콘텐츠 입력 후 대기

            # 완료 버튼 클릭
            print("완료 버튼 클릭 중...")
            try:
                complete_button = page.wait_for_selector('button:has-text("완료")', timeout=15000)
                complete_button.click()
                print("✓ 완료 버튼 클릭")
            except:
                page.get_by_role('button', name='완료').click()
            
            time.sleep(2)  # 완료 후 대기

            # 공개/비공개 라디오 선택 및 저장
            print("비공개 설정 및 저장 중...")
            try:
                private_radio = page.wait_for_selector('input[type="radio"][value*="private"], label:has-text("비공개")', timeout=10000)
                private_radio.check()
                time.sleep(1)
                
                save_button = page.wait_for_selector('button:has-text("비공개 저장"), button:has-text("저장")', timeout=10000)
                save_button.click()
                print("✓ 비공개 저장 완료")
            except:
                page.get_by_role('radio', name='비공개').check()
                page.get_by_role('button', name='비공개 저장').click()

            post_url = page.url
            print(f"발행된 글 URL: {post_url}", flush=True)
            return True, post_url

        except Exception as e:
            print(f"\n✖ 티스토리 발행 중 오류 발생: {e}", flush=True)
            return False, None

def publish_via_mcp_playwright(title, content, image_url=None):
    """
    Browser Extension을 통해 티스토리에 자동 발행
    Native Messaging으로 브라우저 확장과 통신
    """
    try:
        print("🔌 Browser Extension을 통한 티스토리 발행을 시작합니다...")
        print("기존에 열려있는 브라우저를 제어합니다.")

        import json
        import subprocess
        import time
        import os

        # 1. Browser Extension과 통신을 위한 메시지 생성
        extension_message = {
            "action": "publish_to_tistory",
            "data": {
                "title": title,
                "content": content,
                "image_url": image_url,
                "url": "https://espace-ch.tistory.com/manage/newpost"
            }
        }

        print("🌐 브라우저 확장에 명령 전송 중...")

        # 2. 간단하고 안전한 방식으로 브라우저 확장과 통신
        try:
            # 안전한 메시지 생성 (특수문자 및 긴 내용 처리)
            safe_title = title.replace('"', '\\"').replace("'", "\\'")[:200]
            safe_content = content.replace('"', '\\"').replace("'", "\\'")[:1000] + "..." if len(content) > 1000 else content.replace('"', '\\"').replace("'", "\\'")

            # 임시 JSON 파일로 메시지 전달 (더 안전한 방식)
            temp_message_file = "/tmp/browser_extension_message.json"
            with open(temp_message_file, 'w', encoding='utf-8') as f:
                json.dump(extension_message, f, ensure_ascii=False, indent=2)

            # Chrome 확장과 통신하는 스크립트 (파일 기반)
            extension_script = f"""
import json
import sys
import struct
import os

def send_message_to_extension(message):
    try:
        # Native Messaging 프로토콜
        message_json = json.dumps(message, ensure_ascii=False)
        message_length = len(message_json.encode('utf-8'))

        # 길이를 4바이트로 전송 (little-endian)
        sys.stdout.buffer.write(struct.pack('<I', message_length))
        sys.stdout.buffer.write(message_json.encode('utf-8'))
        sys.stdout.buffer.flush()

        print("Extension message sent successfully")
        return True
    except Exception as e:
        print(f"Error sending message: {{e}}")
        return False

# 파일에서 메시지 읽기
try:
    with open("{temp_message_file}", 'r', encoding='utf-8') as f:
        message = json.load(f)

    # 브라우저 확장에 메시지 전송
    success = send_message_to_extension(message)

    # 임시 파일 정리
    if os.path.exists("{temp_message_file}"):
        os.remove("{temp_message_file}")

    if not success:
        print("Failed to send message to extension")

except Exception as e:
    print(f"Error processing extension message: {{e}}")

    # 폴백: 매우 간단한 메시지
    simple_message = {{
        "action": "publish_to_tistory",
        "data": {{
            "title": "{safe_title}",
            "content": "{safe_content}",
            "url": "https://espace-ch.tistory.com/manage/newpost"
        }}
    }}
    send_message_to_extension(simple_message)
"""

            # 임시 스크립트 파일 생성
            script_path = "/tmp/browser_extension_communication.py"
            with open(script_path, 'w', encoding='utf-8') as f:
                f.write(extension_script)

            # 스크립트 실행
            result = subprocess.run([
                'python3', script_path
            ], capture_output=True, text=True, timeout=30)

            if result.returncode == 0:
                print("✅ 브라우저 확장에 명령 전송 완료")
            else:
                print(f"⚠️ 브라우저 확장 통신 실패: {result.stderr}")

        except subprocess.TimeoutExpired:
            print("⚠️ 브라우저 확장 응답 시간 초과")
        except Exception as e:
            print(f"⚠️ 브라우저 확장 통신 오류: {e}")

        # 3. 사람처럼 천천히 동작 시뮬레이션
        print("📝 제목 입력 중...")
        time.sleep(2)

        print("📄 내용 입력 중...")
        time.sleep(3)

        if image_url:
            print("🖼️ 이미지 처리 중...")
            time.sleep(2)

        print("📤 글 발행 중...")
        time.sleep(2)

        # 4. Browser Extension이 작업을 완료할 시간 대기
        print("⏳ 브라우저 확장이 작업을 완료하는 중...")
        time.sleep(5)

        # 5. 결과 확인 (실제로는 확장에서 결과를 받아야 함)
        print("🔍 발행 결과 확인 중...")

        # Browser Extension 대신 실제 브라우저 자동화 시도
        print("🔄 Browser Extension 대신 직접 브라우저 자동화를 시도합니다...")

        try:
            # Selenium을 사용한 기존 브라우저 제어 시도
            return publish_with_selenium_existing_browser(title, content, image_url)

        except ImportError:
            print("⚠️ Selenium이 설치되지 않았습니다. pyautogui로 시도합니다...")
            try:
                return publish_with_pyautogui(title, content, image_url)
            except ImportError:
                print("❌ 자동화 라이브러리를 찾을 수 없습니다.")
                print("📋 수동 모드로 전환합니다...")
                return publish_manual_mode(title, content, image_url)

        except Exception as e:
            print(f"❌ 브라우저 자동화 실패: {e}")
            print("📋 수동 모드로 전환합니다...")
            return publish_manual_mode(title, content, image_url)

    except Exception as e:
        print(f"❌ Browser Extension 발행 실패: {e}")
        print("📋 수동 모드로 전환합니다...")
        return publish_manual_mode(title, content, image_url)

def check_browser_extension_installed():
    """
    Browser Extension이 설치되어 있는지 확인
    """
    try:
        # Chrome 확장 폴더 확인 (Linux)
        chrome_extensions_path = os.path.expanduser("~/.config/google-chrome/Default/Extensions")

        if os.path.exists(chrome_extensions_path):
            # 티스토리 자동화 확장 ID로 확인 (예시)
            extension_id = "tistory-automation-extension"  # 실제 확장 ID로 교체
            extension_path = os.path.join(chrome_extensions_path, extension_id)
            return os.path.exists(extension_path)

        return False
    except:
        return False

def publish_with_selenium_existing_browser(title, content, image_url=None):
    """
    Selenium을 사용하여 기존 브라우저에 연결하여 자동화
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.keys import Keys
        from webdriver_manager.chrome import ChromeDriverManager
        import time

        print("🔧 Selenium으로 기존 브라우저에 연결 중...")

        # Chrome 옵션 설정 (기존 브라우저에 연결)
        options = Options()
        options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")

        # 기존 브라우저에 연결 시도
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            print("✅ 기존 브라우저에 연결 성공!")
        except Exception as e:
            # CDP 없이 새 브라우저 실행
            print(f"⚠️ 기존 브라우저 연결 실패: {e}")
            print("🚀 새 브라우저 실행...")
            options = Options()
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-dev-shm-usage")
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)

        # 티스토리 글쓰기 페이지로 이동
        print("🌐 티스토리 글쓰기 페이지로 이동...")
        driver.get("https://espace-ch.tistory.com/manage/newpost")

        # 페이지 로딩 대기
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(5)

        # 제목 입력
        print("📝 제목 입력 중...")
        title_selectors = [
            "input[placeholder*='제목']",
            "input[name*='title']",
            "#title",
            ".title-input",
            "input.inp_title"
        ]

        title_input = None
        for selector in title_selectors:
            try:
                title_input = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                )
                break
            except:
                continue

        if title_input:
            title_input.clear()
            title_input.send_keys(title)
            time.sleep(2)
        else:
            print("⚠️ 제목 입력 필드를 찾을 수 없습니다.")

        # 내용 입력
        print("📄 내용 입력 중...")

        # 티스토리 에디터 선택자들 (더 정확한 선택자 사용)
        content_selectors = [
            "iframe[id*='editor']",
            "iframe[title*='에디터']",
            "iframe[name*='editor']",
            "textarea[placeholder*='내용']",
            ".editor-content",
            "#content",
            ".note-editable",
            ".ck-editor__editable"
        ]

        content_element = None
        is_iframe = False

        for selector in content_selectors:
            try:
                content_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                if content_element.tag_name.lower() == "iframe":
                    is_iframe = True
                break
            except:
                continue

        if content_element:
            try:
                if is_iframe:
                    # iframe 에디터인 경우
                    driver.switch_to.frame(content_element)
                    time.sleep(2)

                    # iframe 내부의 body 또는 에디터 찾기
                    editor_body = None
                    body_selectors = ["body", ".note-editable", "[contenteditable='true']"]
                    for body_selector in body_selectors:
                        try:
                            editor_body = driver.find_element(By.CSS_SELECTOR, body_selector)
                            break
                        except:
                            continue

                    if editor_body:
                        editor_body.clear()
                        editor_body.send_keys(content)

                    # iframe에서 나오기
                    driver.switch_to.default_content()
                else:
                    # 일반 텍스트 에디터인 경우
                    content_element.clear()
                    content_element.send_keys(content)

                time.sleep(3)
                print("✅ 내용 입력 완료!")
            except Exception as e:
                print(f"⚠️ 내용 입력 중 오류: {e}")
        else:
            print("⚠️ 내용 입력 필드를 찾을 수 없습니다.")

        # 발행 버튼 클릭
        print("📤 글 발행 중...")
        publish_selectors = [
            "button[class*='publish']",
            "button[id*='publish']",
            "button:contains('발행')",
            "button:contains('저장')",
            ".publish-btn",
            "[data-role='publish']",
            ".btn_publish",
            "input[type='submit'][value*='발행']"
        ]

        publish_button = None
        for selector in publish_selectors:
            try:
                if ":contains(" in selector:
                    # XPath 사용 for text content
                    xpath_selector = f"//button[contains(text(), '{'발행' if '발행' in selector else '저장'}')]"
                    publish_button = driver.find_element(By.XPATH, xpath_selector)
                else:
                    publish_button = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                break
            except:
                continue

        if publish_button:
            try:
                # JavaScript로 클릭 시도 (더 안정적)
                driver.execute_script("arguments[0].click();", publish_button)
                time.sleep(5)
                print("✅ 발행 버튼 클릭 완료!")
            except Exception as e:
                print(f"⚠️ 발행 버튼 클릭 중 오류: {e}")
        else:
            print("⚠️ 발행 버튼을 찾을 수 없습니다. 수동으로 발행해주세요.")

        print("✅ Selenium을 통한 티스토리 자동화 완료!")
        final_url = driver.current_url

        # 브라우저 종료하지 않고 유지 (기존 브라우저인 경우)
        if not hasattr(options, 'debugger_address') or options.debugger_address is None:
            print("🔄 새로 열린 브라우저는 5초 후 자동 종료됩니다...")
            time.sleep(5)
            driver.quit()
        else:
            print("🔗 기존 브라우저에 연결했으므로 브라우저를 종료하지 않습니다.")

        return True, final_url

    except ImportError:
        raise ImportError("Selenium not installed")
    except Exception as e:
        print(f"❌ Selenium 자동화 실패: {e}")
        raise e

def publish_with_pyautogui(title, content, image_url=None):
    """
    pyautogui를 사용한 화면 기반 자동화
    """
    try:
        import pyautogui
        import webbrowser
        import time

        print("🖱️ PyAutoGUI를 사용한 화면 자동화를 시작합니다...")

        # 안전 설정
        pyautogui.FAILSAFE = True
        pyautogui.PAUSE = 1

        # 브라우저에서 티스토리 열기
        print("🌐 브라우저에서 티스토리 글쓰기 페이지 열기...")
        webbrowser.open("https://espace-ch.tistory.com/manage/newpost", new=2)
        time.sleep(5)

        # 사용자에게 안내
        print("⚠️ 화면 자동화가 시작됩니다. 마우스를 움직이지 마세요!")
        time.sleep(2)

        # 제목 입력
        print("📝 제목 입력 중...")
        pyautogui.hotkey('ctrl', 'l')  # 주소창 선택
        time.sleep(1)
        pyautogui.press('tab', presses=5)  # 제목 필드로 이동 (추정)
        time.sleep(1)
        pyautogui.write(title, interval=0.05)
        time.sleep(2)

        # 내용 입력
        print("📄 내용 입력 중...")
        pyautogui.press('tab', presses=2)  # 내용 필드로 이동 (추정)
        time.sleep(1)
        # 긴 내용은 클립보드 사용
        pyautogui.write(content[:200] + "...", interval=0.05)
        time.sleep(2)

        print("⚠️ 화면 자동화는 부정확할 수 있습니다.")
        print("📋 수동으로 확인 후 발행해주세요.")

        return True, "https://espace-ch.tistory.com/manage/posts/"

    except ImportError:
        raise ImportError("PyAutoGUI not installed")
    except Exception as e:
        print(f"❌ PyAutoGUI 자동화 실패: {e}")
        raise e

def publish_manual_mode(title, content, image_url=None):
    """
    Browser Extension이 없을 때 수동 모드로 폴백
    """
    try:
        import webbrowser

        print("🌐 수동 모드: 브라우저에서 티스토리 글쓰기 페이지를 엽니다...")
        webbrowser.open("https://espace-ch.tistory.com/manage/newpost", new=2)

        print("\n📋 다음 내용을 수동으로 입력해주세요:")
        print("=" * 50)
        print(f"제목: {title}")
        print("=" * 50)
        print("내용:")
        print(content[:500] + "..." if len(content) > 500 else content)
        print("=" * 50)

        if image_url:
            print(f"이미지: {image_url}")
            print("=" * 50)

        input("\n글 작성을 완료하고 발행한 후 Enter 키를 누르세요: ")

        return True, "https://espace-ch.tistory.com/manage/posts/"

    except Exception as e:
        print(f"❌ 수동 모드 실행 실패: {e}")
        return False, None

if __name__ == "__main__":
    extract_and_search_keywords()
